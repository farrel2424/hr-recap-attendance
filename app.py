"""
Absensi Rekap — Streamlit App
Jalankan dengan: streamlit run app.py
"""

import re
import io
import datetime as _dt
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from database.db import (
    init_db, save_periode, get_periodes, get_rekap,
    get_daily, get_all_daily,
    update_karyawan, update_absensi_row,
    soft_delete_periode,
    get_dates_in_periode,
    get_rules_in_periode,
    bulk_update_h,
    bulk_update_none_corrections,
)
from classifiers import (
    classify,
    classify_str,
    classify_shift_type,
    parse_shift_start,
    parse_shift_end,
    parse_time_to_minutes,
    has_status,
    SKIP_SHIFTS,
    _NOT_PUNCHED,
)

st.set_page_config(
    page_title="Absensi Rekap",
    page_icon="🗓️",
    layout="wide",
    initial_sidebar_state="collapsed",
)
init_db()
if "dialog_target" not in st.session_state:
    st.session_state.dialog_target = None
if "dialog_emp" not in st.session_state:
    st.session_state.dialog_emp = None
if "df_key_suffix" not in st.session_state:
    st.session_state.df_key_suffix = 0
if "current_periode" not in st.session_state:
    st.session_state.current_periode = None
if "show_upload_panel" not in st.session_state:
    st.session_state.show_upload_panel = False
if "show_export_panel" not in st.session_state:
    st.session_state.show_export_panel = False
if "_pending_file_bytes" not in st.session_state:
    st.session_state._pending_file_bytes = None
if "_override_confirmed_for" not in st.session_state:
    st.session_state._override_confirmed_for = None
if "_show_override_confirm" not in st.session_state:
    st.session_state._show_override_confirm = False
if "_pending_override_periode" not in st.session_state:
    st.session_state._pending_override_periode = None
if "show_h_panel" not in st.session_state:
    st.session_state.show_h_panel = False
if "export_prepared" not in st.session_state:
    st.session_state.export_prepared = False
if "_last_export_sel" not in st.session_state:
    st.session_state._last_export_sel = []
if "_force_refresh_cal" not in st.session_state:
    st.session_state._force_refresh_cal = False

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:ital,opsz,wght@0,9..40,300;0,9..40,400;0,9..40,500;0,9..40,600;0,9..40,700;1,9..40,400&family=DM+Mono:wght@400;500&display=swap');

html, body, [class*="css"] { font-family: 'DM Sans', -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; }

:root {
    --bg-primary:      #ffffff;
    --bg-secondary:    #f8fafc;
    --bg-tertiary:     #f1f5f9;
    --border-color:    #e2e8f0;
    --border-strong:   #cbd5e1;
    --text-primary:    #0f172a;
    --text-secondary:  #334155;
    --text-muted:      #64748b;
    --text-faint:      #94a3b8;
    --shadow-sm:       0 1px 3px rgba(0,0,0,0.08);
    --shadow-md:       0 4px 12px rgba(0,0,0,0.10);
    --badge-bg:        #eff6ff;
    --badge-color:     #1d4ed8;
    --table-hover:     #f0f7ff;
    --table-header-bg: #f1f5f9;
    --row-accent-1:    #3b82f6;
    --row-accent-2:    #8b5cf6;
    --row-accent-3:    #0ea5e9;
    --row-accent-4:    #10b981;
}

@media (prefers-color-scheme: dark) {
    :root {
        --bg-primary:      #0f172a;
        --bg-secondary:    #1e293b;
        --bg-tertiary:     #334155;
        --border-color:    #334155;
        --border-strong:   #475569;
        --text-primary:    #f1f5f9;
        --text-secondary:  #cbd5e1;
        --text-muted:      #94a3b8;
        --text-faint:      #64748b;
        --shadow-sm:       0 1px 3px rgba(0,0,0,0.35);
        --shadow-md:       0 4px 12px rgba(0,0,0,0.5);
        --badge-bg:        #1e3a5f;
        --badge-color:     #7dd3fc;
        --table-hover:     #1e3358;
        --table-header-bg: #1e293b;
    }
}
/* Streamlit dark mode class override — hanya aktif di OS dark mode */
@media (prefers-color-scheme: dark) {
    [data-testid="stAppViewContainer"] {
        --bg-primary:      #0f172a;
        --bg-secondary:    #1e293b;
        --text-primary:    #f1f5f9;
        --text-secondary:  #cbd5e1;
        --text-muted:      #94a3b8;
        --text-faint:      #64748b;
        --border-color:    #334155;
        --border-strong:   #475569;
        --table-hover:     #1e3358;
        --table-header-bg: #1e293b;
        --badge-bg:        #1e3a5f;
        --badge-color:     #7dd3fc;
    }
}

.main .block-container { padding: 2rem 3rem 4rem; max-width: 1440px; }

/* ── App Header ── */
.app-header {
    background: linear-gradient(135deg, #0f2027 0%, #203a43 50%, #2c5364 100%);
    border-radius: 20px; padding: 2.5rem 3rem; margin-bottom: 1.5rem;
    position: relative; overflow: hidden;
    box-shadow: 0 8px 32px rgba(0,0,0,0.25);
}
.app-header::before {
    content: ''; position: absolute; top: -50%; right: -10%;
    width: 400px; height: 400px;
    background: radial-gradient(circle, rgba(255,255,255,0.05) 0%, transparent 70%);
    border-radius: 50%;
}
.app-header::after {
    content: ''; position: absolute; bottom: -30%; left: 40%;
    width: 250px; height: 250px;
    background: radial-gradient(circle, rgba(59,130,246,0.08) 0%, transparent 70%);
    border-radius: 50%;
}
.app-header h1 {
    color: #ffffff; font-size: 2.1rem; font-weight: 700;
    margin: 0 0 0.3rem 0; letter-spacing: -0.03em;
}
.app-header p { color: rgba(255,255,255,0.70); font-size: 0.92rem; margin: 0; }
.badge {
    display: inline-block; background: rgba(255,255,255,0.12); color: #7dd3fc;
    padding: 0.2rem 0.75rem; border-radius: 20px; font-size: 0.75rem;
    font-family: 'DM Mono', 'Cascadia Code', 'Fira Mono', 'Courier New', monospace; margin-bottom: 1rem;
    letter-spacing: 0.08em; border: 1px solid rgba(125,211,252,0.25);
}

/* ── Action Panel (Upload / Export) ── */
.action-panel {
    background: var(--bg-secondary);
    border: 1px solid var(--border-color);
    border-radius: 14px;
    padding: 1.4rem 1.8rem;
    margin-bottom: 1.5rem;
    box-shadow: var(--shadow-sm);
    animation: slideDown 0.18s ease-out;
}
@keyframes slideDown {
    from { opacity: 0; transform: translateY(-8px); }
    to   { opacity: 1; transform: translateY(0); }
}
.action-panel-title {
    font-weight: 700; font-size: 0.95rem;
    color: var(--text-primary); margin-bottom: 0.3rem;
    display: flex; align-items: center; gap: 0.5rem;
}
.action-panel-desc {
    font-size: 0.82rem; color: var(--text-muted);
    margin-bottom: 1rem;
}

/* ── Period Table ── */
.period-section-title {
    font-size: 0.78rem; font-weight: 700; color: var(--text-muted);
    text-transform: uppercase; letter-spacing: 0.08em;
    margin-bottom: 1rem; display: flex; align-items: center; gap: 0.5rem;
}
.period-card {
    background: var(--bg-primary);
    border: 1px solid var(--border-color);
    border-radius: 14px; overflow: hidden;
    box-shadow: var(--shadow-sm); margin-bottom: 1.5rem;
}
.period-header {
    display: grid;
    grid-template-columns: 48px 1fr 140px 190px 190px 120px;
    background: var(--table-header-bg);
    border-bottom: 2px solid var(--border-strong);
    padding: 0 1.2rem;
}
.period-header-cell {
    padding: 0.7rem 0.5rem;
    font-size: 0.72rem; font-weight: 700;
    color: var(--text-faint); text-transform: uppercase; letter-spacing: 0.08em;
}
.period-row {
    display: grid;
    grid-template-columns: 48px 1fr 140px 190px 190px 120px;
    padding: 0 1.2rem; border-bottom: 1px solid var(--border-color);
    align-items: center; transition: background 0.12s;
    position: relative;
}
.period-row::before {
    content: ''; position: absolute; left: 0; top: 20%; bottom: 20%;
    width: 3px; border-radius: 0 2px 2px 0; opacity: 0.6;
    background: var(--row-accent);
}
.period-row:last-child { border-bottom: none; }
.period-row:hover { background: var(--table-hover); }
.period-row:hover::before { opacity: 1; }
.period-cell { padding: 0.9rem 0.5rem; font-size: 0.86rem; color: var(--text-secondary); }
.period-cell.no {
    display: flex; align-items: center; justify-content: center;
    font-size: 0.78rem; color: var(--text-faint);
    font-family: 'DM Mono', monospace;
}
.period-cell.month { font-weight: 700; color: var(--text-primary); font-size: 0.92rem; }
.period-badge {
    display: inline-flex; align-items: center;
    background: var(--badge-bg); color: var(--badge-color);
    padding: 0.22rem 0.7rem; border-radius: 20px;
    font-family: 'DM Mono', monospace; font-size: 0.73rem;
    font-weight: 600; border: 1px solid rgba(29, 78, 216, 0.30);
}
@media (prefers-color-scheme: dark) {
    .period-badge { border-color: rgba(125, 211, 252, 0.30); }
}
.empty-state {
    text-align: center; padding: 5rem 2rem; color: var(--text-faint);
}
.empty-state .icon { font-size: 4rem; margin-bottom: 1.2rem; opacity: 0.6; }
.empty-state .title {
    font-size: 1.1rem; font-weight: 700;
    color: var(--text-muted); margin-bottom: 0.5rem;
}
.empty-state .subtitle { font-size: 0.88rem; }

/* ── Metric Cards ── */
.metric-row {
    display: grid; grid-template-columns: repeat(auto-fill, minmax(160px, 1fr));
    gap: 1rem; margin: 2rem 0 2.5rem;
}
.metric-card {
    border-radius: 14px; padding: 1.4rem 1.6rem 1.3rem;
    font-weight: 600; position: relative; overflow: hidden;
    transition: transform 0.15s, box-shadow 0.15s;
}
.metric-card:hover { transform: translateY(-2px); box-shadow: var(--shadow-md); }
.metric-card::after {
    content: ''; position: absolute; bottom: -20px; right: -20px;
    width: 80px; height: 80px; border-radius: 50%; opacity: 0.06; background: currentColor;
}
.metric-shift    { background: #f0fdf4; border-left: 4px solid #22c55e; }
.metric-late     { background: #fffbeb; border-left: 4px solid #f59e0b; }
.metric-k        { background: #fef2f2; border-left: 4px solid #ef4444; }
.metric-total    { background: #eff6ff; border-left: 4px solid #3b82f6; }
.metric-al       { background: #fdf4ff; border-left: 4px solid #a855f7; }
.metric-half-al  { background: #fff1f2; border-left: 4px solid #fb7185; }
.metric-wfa      { background: #f0f9ff; border-left: 4px solid #0ea5e9; }
.metric-half-wfa { background: #eff6ff; border-left: 4px solid #60a5fa; }
.metric-wfs      { background: #eef2ff; border-left: 4px solid #6366f1; }
.metric-dw       { background: #fff7ed; border-left: 4px solid #f97316; }
.metric-ksick    { background: #fdf2f8; border-left: 4px solid #ec4899; }
.metric-off      { background: #f8fafc; border-left: 4px solid #94a3b8; }
.metric-ul       { background: #f0fdfa; border-left: 4px solid #14b8a6; }
.metric-hl       { background: #fffbeb; border-left: 4px solid #eab308; }
.metric-ml       { background: #f0fdf4; border-left: 4px solid #4ade80; }
.metric-wml      { background: #f0f9ff; border-left: 4px solid #22d3ee; }
.metric-ot       { background: #f8fafc; border-left: 4px solid #94a3b8; }
.metric-rl       { background: #f0fdf4; border-left: 4px solid #4ade80; }
.metric-h        { background: #fff0f0; border-left: 4px solid #ff4444; }

.metric-card .label {
    font-size: 0.72rem; color: var(--text-muted); font-weight: 600;
    text-transform: uppercase; letter-spacing: 0.08em; margin-bottom: 0.5rem;
    display: flex; align-items: center; gap: 0.35rem;
}

.metric-card .value {
    font-size: 2.1rem; font-weight: 700;
    font-family: 'DM Mono', monospace; line-height: 1;
}
.metric-shift    .value { color: #16a34a; }
.metric-late     .value { color: #d97706; }
.metric-k        .value { color: #dc2626; }
.metric-total    .value { color: #2563eb; }
.metric-al       .value { color: #9333ea; }
.metric-half-al  .value { color: #e11d48; }
.metric-wfa      .value { color: #0284c7; }
.metric-half-wfa .value { color: #2563eb; }
.metric-wfs      .value { color: #4338ca; }
.metric-dw       .value { color: #ea580c; }
.metric-ksick    .value { color: #db2777; }
.metric-off      .value { color: #64748b; }
.metric-ul       .value { color: #0f766e; }
.metric-hl       .value { color: #a16207; }
.metric-ml       .value { color: #15803d; }
.metric-wml      .value { color: #0e7490; }
.metric-ot       .value { color: #475569; }
.metric-rl       .value { color: #15803d; }
.metric-h        .value { color: #cc0000; }
.metric-card .sub { font-size: 0.70rem; color: var(--text-faint); font-weight: 400; margin-top: 0.35rem; }

/* ── Download Button ── */
.stDownloadButton button {
    background: linear-gradient(135deg, #1e40af, #3b82f6) !important;
    color: white !important; border: none !important;
    padding: 0.6rem 1.8rem !important; border-radius: 8px !important;
    font-weight: 600 !important; font-size: 0.9rem !important;
    transition: all 0.2s !important;
}
.stDownloadButton button:hover {
    transform: translateY(-1px) !important;
    box-shadow: 0 4px 12px rgba(59,130,246,0.4) !important;
}

.section-title {
    font-size: 1rem; font-weight: 700; color: var(--text-primary);
    letter-spacing: -0.01em; margin: 0 0 1rem 0;
    display: flex; align-items: center; gap: 0.5rem;
}
.streamlit-expanderHeader { font-weight: 600 !important; }
#MainMenu, footer { visibility: hidden; }

/* ── Responsive Breakpoints ── */
@media (max-width: 900px) {
    .main .block-container { padding: 1rem 1.2rem 3rem; }
    .app-header { padding: 1.5rem 1.5rem; }
    .app-header h1 { font-size: 1.5rem; }
    .metric-card .value { font-size: 1.7rem; }
}
@media (max-width: 600px) {
    .main .block-container { padding: 0.75rem 0.75rem 2rem; }
    .metric-card .value { font-size: 1.4rem; }
    .metric-card { padding: 1rem 1rem 0.9rem; }
}

/* ── Sticky Column — override st.dataframe internal ── */
/* Freeze kolom No. dan Nama di st.dataframe */
[data-testid="stDataFrame"] [data-testid="glideDataEditor"] .dvn-scroller {
    overflow-x: auto !important;
}
/* Header sticky row */
[data-testid="stDataFrame"] canvas {
    /* glide-data-grid menggunakan canvas — tidak bisa di-CSS */
}

/* Wrapper: beri tinggi tetap agar scroll vertikal aktif */
[data-testid="stDataFrame"] > div {
    border-radius: 12px;
    overflow: hidden;
    border: 1px solid var(--border-color);
}

/* ── Keyboard Accessibility: :focus-visible ── */
[data-testid="stButton"] > button:focus-visible,
[data-testid="stDownloadButton"] > button:focus-visible,
[data-testid="stFormSubmitButton"] > button:focus-visible {
    outline: 2px solid #3b82f6 !important;
    outline-offset: 2px !important;
    box-shadow: 0 0 0 4px rgba(59,130,246,0.18) !important;
}
[data-testid="stSelectbox"] > div:focus-visible,
[data-testid="stMultiSelect"] > div:focus-visible {
    outline: 2px solid #3b82f6;
    outline-offset: 2px;
    border-radius: 8px;
}
[data-testid="stCheckbox"] input:focus-visible ~ div,
[data-testid="stRadio"] input:focus-visible ~ div {
    outline: 2px solid #3b82f6;
    outline-offset: 2px;
    border-radius: 4px;
}

/* ── Sub-table classes ── */
.nt-sub-wrap {
    margin: 0 0 0.4rem 1.5rem;
    border: 1px solid var(--border-color, #e2e8f0);
    border-top: 2px solid rgba(245,158,11,.35);
    border-radius: 0 0 8px 8px;
    overflow: hidden;
}
.nt-sub-head-row { background: #fffbeb; }
.nt-sub-row-even { background: #f8fafc; }
.nt-sub-row-odd  { background: #ffffff; }
.nt-sub-date {
    padding: .4rem .8rem; font-family: monospace;
    font-size: .78rem; color: #475569; white-space: nowrap;
}
.nt-sub-reason { padding: .4rem .8rem; }
@media (prefers-color-scheme: dark) {
    .nt-sub-head-row { background: #1c1917 !important; }
    .nt-sub-row-even { background: #1e293b !important; }
    .nt-sub-row-odd  { background: #0f172a !important; }
    .nt-sub-date     { color: #94a3b8 !important; }
}

</style>
""", unsafe_allow_html=True)



def _find_col(df: pd.DataFrame, prefix: str) -> str | None:
    for col in df.columns:
        if str(col).startswith(prefix):
            return col
    return None

def _find_ksick_col(df)         -> str | None: return _find_col(df, "K-Sick W Letter")
def _find_al_col(df)            -> str | None: return _find_col(df, "AnnualLeave")
def _find_ul_col(df)            -> str | None: return _find_col(df, "UL-Unpaid")
def _find_duration_late_col(df) -> str | None: return _find_col(df, "Duration of late arrival")
def _find_duration_early_col(df)-> str | None: return _find_col(df, "Duration of early departure")
def _find_wfh_col(df)           -> str | None: return _find_col(df, "WFH-WorkFromHome")
def _find_offsite_col(df)       -> str | None: return _find_col(df, "Offsite(Hour)")
def _find_missed_punch_col(df)  -> str | None: return _find_col(df, "Number of missed punches")
def _find_hl_col(df)            -> str | None: return _find_col(df, "HL-Happy")
def _find_ml_col(df)            -> str | None: return _find_col(df, "ML-Maternity")
def _find_wml_col(df)           -> str | None: return _find_col(df, "WML-WifeMater")
def _find_ot_col(df)            -> str | None: return _find_col(df, "OT - Others")
def _find_rl_col(df) -> str | None: return _find_col(df, "RL - Roster Leave")
def _find_pl_col(df) -> str | None: return _find_col(df, "PL-Personal(TKA)")

def _get_sheet_name(file_bytes: bytes) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheets = wb.sheetnames
    wb.close()
    for candidate in ("General statistics and attendan",
                      "General statistics and attendance details"):
        if candidate in sheets:
            return candidate
    raise ValueError(
        "Sheet 'General statistics and attendan' atau "
        "'General statistics and attendance details' tidak ditemukan."
    )

_STATUS_ICON = {
    "S"      : "📋",
    "Late"   : "🕐",
    "1/2 UL" : "⛔",
    "UL"     : "📋",
    "AL"     : "🌴",
    "1/2 AL" : "🌗",
    "WFA"    : "🏠",
    "1/2 WFA": "🏡",
    "WFS"    : "📍",
    "DW"     : "🚫",
    "K"      : "💊",
    "Off"    : "🏖️",
    "HL"     : "💍",
    "ML"     : "🤱",
    "WML"    : "👶",
    "OT"     : "📝",
    "RL"     : "📅",
    "H"      : "🔴",
    "PL"     : "🪪",
    "None"   : "❓",
}

def _fmt_klasifikasi(klas_raw) -> str:
    if not klas_raw:
        return "❓ None"
    parts = []
    for s in klas_raw:
        icon = _STATUS_ICON.get(s, "❓")
        parts.append(f"{icon} {s}")
    return " / ".join(parts)


# ──────────────────────────────────────────────────────────────
# Helper: Ekspor Kalender — Label & Warna Sel
# ──────────────────────────────────────────────────────────────

# Pemetaan label teks untuk tiap klasifikasi di ekspor kalender
_LABEL_MAP = {
    "Off":     "OFF",
    "AL":      "AL",
    "1/2 AL":  "0,5AL",
    "WFA":     "WFA",
    "1/2 WFA": "0,5WFA",
    "WFS":     "WFS",
    "K":       "K",
    "DW":      "DW",
    "UL":      "UL",
    "1/2 UL":  "0,5UL",
    "Late":    "L",
    "S":       "S",
    "HL":      "HL",
    "ML":      "ML",
    "WML":     "WML",
    "OT":      "OT",
    "RL":      "RL",
    "H":       "H",
    "PL":      "PL",
}


# Pemetaan background color (hex tanpa '#') untuk tiap label ekspor
# Palet warna pastel — mudah dibaca, konsisten dengan tema aplikasi
_CELL_FILL: dict[str, PatternFill] = {
    "S":      PatternFill("solid", fgColor="D9EAD3"),  # hijau muda  — hadir normal
    "OFF":    PatternFill("solid", fgColor="CFE2F3"),  # biru muda   — Off / libur
    "AL":     PatternFill("solid", fgColor="D9D2E9"),  # ungu muda   — Annual Leave penuh
    "0,5AL":  PatternFill("solid", fgColor="EAD1DC"),  # pink        — Half Annual Leave
    "WFA":    PatternFill("solid", fgColor="D0E4F7"),  # biru langit — WFA penuh
    "0,5WFA": PatternFill("solid", fgColor="C9DAF8"),  # biru medium — Half WFA
    "WFS":    PatternFill("solid", fgColor="B4C7E7"),  # biru indigo — Work From Offsite
    "UL":     PatternFill("solid", fgColor="B7E1CD"),  # hijau sedang— Unpaid Leave penuh
    "0,5UL":  PatternFill("solid", fgColor="FCE5CD"),  # oranye muda — Half UL / telat >120 mnt
    "L":      PatternFill("solid", fgColor="FFF2CC"),  # kuning      — Late / pulang cepat
    "K":      PatternFill("solid", fgColor="F4CCCC"),  # merah muda  — Sakit dengan surat
    "DW":     PatternFill("solid", fgColor="EA9999"),  # merah       — Absence / tidak hadir
    "HL":     PatternFill("solid", fgColor="FFE599"),  # kuning emas  — cuti pernikahan
    "ML":     PatternFill("solid", fgColor="B6D7A8"),  # hijau sedang — cuti melahirkan
    "WML":    PatternFill("solid", fgColor="A2C4C9"),  # teal         — cuti istri melahirkan
    "OT":     PatternFill("solid", fgColor="D9D9D9"),  # abu-abu
    "RL":     PatternFill("solid", fgColor="D9EAD3"),  # hijau muda — roster leave      — cuti lainnya
    "H":      PatternFill("solid", fgColor="FF9999"),  # merah cerah — hari libur nasional
    "PL":     PatternFill("solid", fgColor="E6D0DE"),  # ungu muda kemerahan — Personal Leave TKA
}



def determine_reason(
    shift: str,
    att_result: str,
    status_klasifikasi,
    jam_masuk: str = "",
    jam_keluar: str = "",
) -> str:
    """
    Tentukan alasan mengapa sebuah baris absensi tidak menghasilkan klasifikasi
    (sel kalender kosong / putih).

    Urutan pengecekan:
      1. att_result kosong / NaN / "--"
           → "Att result tidak tercatat"
      2. shift masuk SKIP_SHIFTS (Rest / Not scheduled / "--" / kosong)
           → "Shift '{nilai}' dilewati engine"
      3. status_klasifikasi None / NULL, tapi shift & att valid
           → "Att result tidak dikenali: '{nilai}'"
      4. Fallback
           → "Tidak dapat ditentukan"
    """
    _att = str(att_result).strip() if att_result else ""
    _shift = str(shift).strip() if shift else ""

    # 1. Att result tidak ada sama sekali
    if not _att or _att in {"--", "nan", "None"}:
        return "Att result tidak tercatat"

    # 2. Shift dilewati engine
    if _shift in SKIP_SHIFTS:
        _label = f"'{_shift}'" if _shift else "(kosong)"
        return f"Shift {_label} dilewati"

    # 3. Shift & att valid tapi tidak cocok dengan pola manapun
    if not status_klasifikasi or str(status_klasifikasi).strip() in {"", "None", "nan"}:
        return f"Att result tidak dikenali: '{_att}'"

    # 4. Fallback
    return "Tidak dapat ditentukan"

def _get_cell_display(shift_text: str, classification) -> str:
    """Tentukan nilai label teks untuk sel kalender berdasarkan klasifikasi."""
    return _LABEL_MAP.get(classification, "")


def parse_date_from_time(val):
    if not isinstance(val, str):
        return str(val) if val else ""
    m = re.search(r'(\d{4}/\d{2}/\d{2})', val)
    return m.group(1) if m else val.strip()


@st.cache_data(show_spinner=False)
def get_employee_daily(file_bytes, account):
    buf = io.BytesIO(file_bytes)
    df_all = pd.read_excel(
        buf,
        sheet_name=_get_sheet_name(file_bytes),
        header=3,
        dtype={"Earliest": str, "Latest": str},
    ).rename(columns={"Unnamed: 0": "Time_Date", "Unnamed: 1": "Name", "Unnamed: 2": "Account"})
    df_emp = df_all[df_all["Account"].astype(str).str.strip() == account].copy()

    k_sick_col      = _find_ksick_col(df_emp)
    al_col          = _find_al_col(df_emp)
    ul_col          = _find_ul_col(df_emp)
    dur_late_col    = _find_duration_late_col(df_emp)
    dur_early_col   = _find_duration_early_col(df_emp)
    wfh_col         = _find_wfh_col(df_emp)
    offsite_col     = _find_offsite_col(df_emp)
    missed_punch_col  = _find_missed_punch_col(df_emp)
    hl_col            = _find_hl_col(df_emp)
    ml_col            = _find_ml_col(df_emp)
    wml_col           = _find_wml_col(df_emp)
    ot_col            = _find_ot_col(df_emp)
    rl_col            = _find_rl_col(df_emp)
    pl_col            = _find_pl_col(df_emp)   # atau df_all / df

    def _parse_hours(val):
        if val is None:
            return 0.0
        if isinstance(val, (int, float)):
            return 0.0 if pd.isna(val) else float(val)
        try:
            return float(str(val).strip())
        except Exception:
            return 0.0

    df_emp["Tanggal"]   = df_emp["Time_Date"].astype(str).apply(parse_date_from_time)
    df_emp["Jam Kerja"] = df_emp["Actual working hours(Hour)"].apply(_parse_hours)

    rows = []
    for _, r in df_emp.iterrows():
        shift_clean = str(r["Shift"]).strip() if isinstance(r["Shift"], str) else ""

        _klas_raw = classify(
            r["Earliest"], r["Shift"], r["Attendance results"],
            latest_raw=r["Latest"],
            leave_app=r.get("Leave & Overtime Application"),
            absences_count=r.get("Number of absences(Count)"),
            k_sick_count=r.get(k_sick_col)     if k_sick_col    else None,
            al_count=r.get(al_col)              if al_col        else None,
            ul_count=r.get(ul_col)              if ul_col        else None,
            duration_late=r.get(dur_late_col)   if dur_late_col  else None,
            duration_early=r.get(dur_early_col) if dur_early_col else None,
            wfh_count=r.get(wfh_col)            if wfh_col       else None,
            offsite_hour=r.get(offsite_col)     if offsite_col   else None,
            missed_punch_count=r.get(missed_punch_col) if missed_punch_col else None,
            hl_count=r.get(hl_col)   if hl_col   else None,
            ml_count=r.get(ml_col)   if ml_col   else None,
            wml_count=r.get(wml_col) if wml_col  else None,
            ot_count=r.get(ot_col)   if ot_col   else None,
            rl_count=r.get(rl_col)   if rl_col   else None,
            pl_count=r.get(pl_col)   if pl_col   else None,
        )

        if _klas_raw is None:
            _klas_raw = ["None"]

        _klas_display = _fmt_klasifikasi(_klas_raw)

        rows.append({
            "Tanggal"        : r["Tanggal"],
            "Shift"          : shift_clean,
            "Jam Masuk"      : str(r["Earliest"]).strip() if pd.notna(r["Earliest"]) else "--",
            "Jam Keluar"     : str(r["Latest"]).strip()   if pd.notna(r["Latest"])   else "--",
            "Status"         : str(r["Attendance results"]).strip() if pd.notna(r["Attendance results"]) else "--",
            "Jam Kerja"      : r["Jam Kerja"],
            "Klasifikasi"    : _klas_display,
            "Klasifikasi_raw": _klas_raw,
        })

    detail_df = pd.DataFrame(rows).sort_values("Tanggal").reset_index(drop=True)
    detail_df.insert(0, "No.", range(1, len(detail_df) + 1))

    n_shift   = int(detail_df["Klasifikasi_raw"].apply(lambda x: has_status(x, "S")).sum())
    n_off     = int(detail_df["Klasifikasi_raw"].apply(lambda x: has_status(x, "Off")).sum())
    jam_shift = float(detail_df[detail_df["Klasifikasi_raw"].apply(lambda x: has_status(x, "S"))]["Jam Kerja"].sum())
    jam_off   = float(detail_df[detail_df["Klasifikasi_raw"].apply(lambda x: has_status(x, "Off"))]["Jam Kerja"].sum())
    summary_df = pd.DataFrame([
        {"Kategori": "S",   "Hari": n_shift, "Total_Jam": jam_shift},
        {"Kategori": "Off", "Hari": n_off,   "Total_Jam": jam_off},
    ])
    return detail_df, summary_df


@st.cache_data(show_spinner=False)
def get_employee_daily_from_db(account, periode):
    df_db = get_daily(account, periode)
    if df_db.empty:
        return pd.DataFrame(), pd.DataFrame()

    rows = []
    for _, r in df_db.iterrows():
        klas_str  = str(r.get("status_klasifikasi") or "").strip()
        if "|" in klas_str:
            klas_raw = [s.strip() for s in klas_str.split("|")]
        elif klas_str:
            _tmp = klas_str.replace("1/2", "\x00HALF\x00")
            _parts = [p.replace("\x00HALF\x00", "1/2").strip() for p in _tmp.split("/")]
            klas_raw = [p for p in _parts if p]
        else:
            klas_raw = ["None"]

        klas_disp = _fmt_klasifikasi(klas_raw)

        jam_masuk  = str(r.get("jam_masuk")  or "").strip() or "--"
        jam_keluar = str(r.get("jam_keluar") or "").strip() or "--"
        status_ab  = str(r.get("status_absensi") or "").strip() or "--"

        rows.append({
            "Tanggal"         : r["tanggal"],
            "Shift"           : str(r.get("shift") or "").strip() or "--",
            "Jam Masuk"       : jam_masuk,
            "Jam Keluar"      : jam_keluar,
            "Status"          : status_ab,
            "Jam Kerja"       : float(r.get("jam_kerja") or 0),
            "Klasifikasi"     : klas_disp,
            "Klasifikasi_raw" : klas_raw,
            "Catatan"         : str(r.get("catatan") or "").strip(),
            "Manual_Override" : bool(r.get("is_manual_override") or 0),
        })

    if not rows:
        return pd.DataFrame(), pd.DataFrame()

    detail_df = pd.DataFrame(rows).sort_values("Tanggal").reset_index(drop=True)
    detail_df.insert(0, "No.", range(1, len(detail_df) + 1))

    n_shift   = int(detail_df["Klasifikasi_raw"].apply(lambda x: has_status(x, "S")).sum())
    n_off     = int(detail_df["Klasifikasi_raw"].apply(lambda x: has_status(x, "Off")).sum())
    jam_shift = float(detail_df[detail_df["Klasifikasi_raw"].apply(lambda x: has_status(x, "S"))]["Jam Kerja"].sum())
    jam_off   = float(detail_df[detail_df["Klasifikasi_raw"].apply(lambda x: has_status(x, "Off"))]["Jam Kerja"].sum())
    summary_df = pd.DataFrame([
        {"Kategori": "S",   "Hari": n_shift, "Total_Jam": jam_shift},
        {"Kategori": "Off", "Hari": n_off,   "Total_Jam": jam_off},
    ])
    return detail_df, summary_df


@st.cache_data(show_spinner=False)
def get_all_daily_for_calendar(file_bytes):
    buf = io.BytesIO(file_bytes)
    df_all = pd.read_excel(
        buf,
        sheet_name=_get_sheet_name(file_bytes),
        header=3,
        dtype={"Earliest": str, "Latest": str},
    ).rename(columns={"Unnamed: 0": "Time_Date", "Unnamed: 1": "Name", "Unnamed: 2": "Account"})

    k_sick_col    = _find_ksick_col(df_all)
    al_col        = _find_al_col(df_all)
    ul_col        = _find_ul_col(df_all)
    dur_late_col  = _find_duration_late_col(df_all)
    dur_early_col = _find_duration_early_col(df_all)
    wfh_col       = _find_wfh_col(df_all)
    offsite_col   = _find_offsite_col(df_all)
    missed_punch_col  = _find_missed_punch_col(df_all)

    hl_col            = _find_hl_col(df_all)
    ml_col            = _find_ml_col(df_all)
    wml_col           = _find_wml_col(df_all)
    ot_col            = _find_ot_col(df_all)
    rl_col            = _find_rl_col(df_all)
    pl_col            = _find_pl_col(df_all)

    df_all = df_all[df_all["Account"].notna() & df_all["Rules"].notna()]
    df_all = df_all[~df_all["Account"].astype(str).str.strip().isin(["", "--"])]

    rows = []
    for _, r in df_all.iterrows():
        account = str(r["Account"]).strip()
        name    = str(r["Name"]).strip() if pd.notna(r.get("Name")) else ""
        shift_t = str(r["Shift"]).strip() if isinstance(r["Shift"], str) else ""

        raw_time = str(r.get("Time_Date", ""))
        m = re.search(r'(\d{4})/(\d{2})/(\d{2})', raw_time)
        if not m:
            continue
        date_str = f"{m.group(1)}-{m.group(2)}-{m.group(3)}"

        klas_raw = classify(
            r["Earliest"], r["Shift"], r["Attendance results"],
            latest_raw=r["Latest"],
            leave_app=r.get("Leave & Overtime Application"),
            absences_count=r.get("Number of absences(Count)"),
            k_sick_count=r.get(k_sick_col)     if k_sick_col    else None,
            al_count=r.get(al_col)              if al_col        else None,
            ul_count=r.get(ul_col)              if ul_col        else None,
            duration_late=r.get(dur_late_col)   if dur_late_col  else None,
            duration_early=r.get(dur_early_col) if dur_early_col else None,
            wfh_count=r.get(wfh_col)            if wfh_col       else None,
            offsite_hour=r.get(offsite_col)     if offsite_col   else None,
            missed_punch_count=r.get(missed_punch_col) if missed_punch_col else None,
            hl_count=r.get(hl_col)   if hl_col   else None,
            ml_count=r.get(ml_col)   if ml_col   else None,
            wml_count=r.get(wml_col) if wml_col  else None,
            ot_count=r.get(ot_col)   if ot_col   else None,
            rl_count=r.get(rl_col)   if rl_col   else None,
        )
        classification = klas_raw[0] if klas_raw else None

        rows.append({
            "Account":        account,
            "Name":           name,
            "Date":           date_str,
            "Shift":          shift_t,
            "Classification": classification,
        })
    return pd.DataFrame(rows)


def _get_all_daily_from_db(periode):
    if not periode:
        return pd.DataFrame(columns=["Account", "Name", "Date", "Shift", "Classification"])
    try:
        df_db = get_all_daily(periode)
    except Exception:
        return pd.DataFrame(columns=["Account", "Name", "Date", "Shift", "Classification"])
    if df_db.empty:
        return pd.DataFrame(columns=["Account", "Name", "Date", "Shift", "Classification"])

    rows = []
    for _, r in df_db.iterrows():
        klas_str = str(r.get("status_klasifikasi") or "").strip()
        if "|" in klas_str:
            parts = [p.strip() for p in klas_str.split("|")]
            klas = parts[0] if parts else None
        else:
            klas = klas_str if klas_str else None

        rows.append({
            "Account":        str(r["account"]).strip(),
            "Name":           str(r["nama"]).strip(),
            "Date":           str(r["tanggal"]).strip(),
            "Shift":          str(r.get("shift") or "").strip(),
            "TipeShift":      str(r.get("tipe_shift") or "").strip(),
            "JamMasuk":       str(r.get("jam_masuk") or "").strip(),
            "JamKeluar":      str(r.get("jam_keluar") or "").strip(),
            "AttResult":      str(r.get("status_absensi") or "").strip(),
            "Classification": klas,
            "Remarks":        str(r.get("catatan") or "").strip(),
        })
    return pd.DataFrame(rows)


@st.dialog("📋 Rincian Harian Karyawan", width="large", dismissible=False)
def show_daily_detail(account, nama, rules, file_bytes=None, periode=None):
    with st.spinner("⏳ Memuat rincian harian..."):
        if periode is not None:
            detail_df, summary_df = get_employee_daily_from_db(account, periode)
        elif file_bytes is not None:
            detail_df, summary_df = get_employee_daily(file_bytes, account)
        else:
            st.error("Tidak ada data yang bisa dimuat.")
            return

    if detail_df.empty:
        st.warning("⚠️ Tidak ada data harian ditemukan untuk karyawan ini.")
        return

    source_label = "📂 Dari file Excel" if file_bytes is not None else "🗄️ Dari database"
    st.markdown(
        '<div style="background:#f8fafc;border-radius:10px;padding:1rem 1.4rem;'
        'border-left:4px solid #3b82f6;margin-bottom:1.2rem;">'
        f'<div style="font-size:1.05rem;font-weight:700;color:#1e293b;">👤 {nama}</div>'
        f'<div style="font-size:0.82rem;color:#64748b;margin-top:0.2rem;">'
        f'<code>{account}</code> &nbsp;·&nbsp; 📌 {rules} &nbsp;·&nbsp; '
        f'<span style="color:#94a3b8">{source_label}'
        + (f" — 📅 Periode {periode}" if periode else "") +
        '</span></div></div>',
        unsafe_allow_html=True,
    )

    tipe_cfg = {
        "S"  : ("☀️ Shift", "#f0fdf4", "#22c55e", "#166534"),
        "Off": ("🏖️ Off",   "#fff7ed", "#fb923c", "#9a3412"),
    }
    cols = st.columns(2)
    for i, key in enumerate(["S", "Off"]):
        row = summary_df[summary_df["Kategori"] == key]
        hari = int(row["Hari"].values[0])        if len(row) else 0
        jam  = float(row["Total_Jam"].values[0]) if len(row) else 0.0
        label, bg, border_c, text_c = tipe_cfg[key]
        with cols[i]:
            st.markdown(
                f'<div style="background:{bg};border-left:4px solid {border_c};'
                f'border-radius:10px;padding:1rem 1.2rem;text-align:center;">'
                f'<div style="font-size:0.75rem;color:#64748b;font-weight:500;'
                f'text-transform:uppercase;letter-spacing:.06em;">{label}</div>'
                f'<div style="font-size:1.9rem;font-weight:700;color:{text_c};">'
                f'{hari}<span style="font-size:1rem"> hari</span></div>'
                f'<div style="font-size:0.8rem;color:#64748b;margin-top:0.1rem;">'
                f'⏱️ {jam:.1f} jam kerja</div></div>',
                unsafe_allow_html=True,
            )

    st.markdown("<div style='margin-top:1.2rem'></div>", unsafe_allow_html=True)

    # ── Edit Data Karyawan & Absensi Harian ──────────────────────────────
    if periode is not None:
        st.markdown(
            "<hr style='border-color:#e2e8f0;margin:1.5rem 0 1rem'>",
            unsafe_allow_html=True,
        )
        with st.expander("✏️ Edit Data Karyawan & Absensi Harian", expanded=False):
            _edit_tab_emp, _edit_tab_abs = st.tabs(["👤 Edit Karyawan", "📅 Edit Absensi Harian"])

            # ── Tab 1: Edit data karyawan ─────────────────────────────
            with _edit_tab_emp:
                st.markdown(
                    '<div style="font-size:0.82rem;color:#64748b;margin-bottom:0.8rem;">'
                    'Perubahan disimpan ke database dan tercermin di rekap berikutnya.</div>',
                    unsafe_allow_html=True,
                )
                with st.form(key=f"_form_emp_{account}_{periode}"):
                    _ec1, _ec2 = st.columns(2)
                    with _ec1:
                        _new_nama  = st.text_input("Nama Karyawan", value=nama)
                    with _ec2:
                        _new_rules = st.text_input("Rules / Departemen", value=rules)
                    _submitted_emp = st.form_submit_button("💾 Simpan Data Karyawan", type="primary")
                if _submitted_emp:
                    try:
                        update_karyawan(account, _new_nama, _new_rules)
                        st.cache_data.clear()
                        st.success(f"✅ Data karyawan **{_new_nama}** berhasil diperbarui.")
                    except Exception as _e_kary:
                        st.error(f"❌ Gagal menyimpan: {_e_kary}")

            # ── Tab 2: Edit absensi harian ────────────────────────────
            with _edit_tab_abs:
                _ALL_STATUS_OPTS = [
                    "S", "Late", "1/2 UL", "UL", "AL", "1/2 AL",
                    "WFA", "1/2 WFA", "WFS", "DW", "K", "Off",
                    "HL", "ML", "WML", "OT", "RL", "H", "PL", "None",
                ]
                st.markdown(    
                    '<div style="font-size:0.82rem;color:#64748b;margin-bottom:0.6rem;">'
                    'Edit jam masuk/keluar, pilih klasifikasi manual, dan tambah catatan. '
                    'Centang <b>Override</b> untuk menandai baris yang ditetapkan manual '
                    '(bukan hasil engine otomatis).</div>',
                    unsafe_allow_html=True,
                )

                _raw_daily = get_daily(account, periode)
                if _raw_daily.empty:
                    st.info("Tidak ada data harian yang dapat diedit.")
                else:
                    _erows = []
                    for _, _er in _raw_daily.iterrows():
                        _ks = str(_er.get("status_klasifikasi") or "").strip()
                        _kfirst = (_ks.split("|")[0].strip() if "|" in _ks else _ks) or "None"
                        if _kfirst not in _ALL_STATUS_OPTS:
                            _kfirst = "None"
                        _erows.append({
                            "Tanggal"    : str(_er["tanggal"]),
                            "Shift"      : str(_er.get("shift") or ""),
                            "Jam Masuk"  : str(_er.get("jam_masuk")  or ""),
                            "Jam Keluar" : str(_er.get("jam_keluar") or ""),
                            "Klasifikasi": _kfirst,
                            "Catatan"    : str(_er.get("catatan")    or ""),
                            "Override"   : bool(_er.get("is_manual_override") or 0),
                        })
                    _edit_df = pd.DataFrame(_erows)

                    _edited_df = st.data_editor(
                        _edit_df,
                        key=f"_editor_{account}_{periode}",
                        use_container_width=True,
                        height=min(60 + len(_edit_df) * 35, 440),
                        hide_index=True,
                        column_config={
                            "Tanggal"    : st.column_config.TextColumn(
                                "📅 Tanggal", disabled=True, width="medium"),
                            "Shift"      : st.column_config.TextColumn(
                                "⏰ Shift", disabled=True, width="medium"),
                            "Jam Masuk"  : st.column_config.TextColumn(
                                "🕐 Jam Masuk", width="small"),
                            "Jam Keluar" : st.column_config.TextColumn(
                                "🕔 Jam Keluar", width="small"),
                            "Klasifikasi": st.column_config.SelectboxColumn(
                                "📊 Klasifikasi",
                                options=_ALL_STATUS_OPTS,
                                required=True,
                                width="small",
                            ),
                            "Catatan"    : st.column_config.TextColumn(
                                "📝 Catatan / Keterangan", width="large"),
                            "Override"   : st.column_config.CheckboxColumn(
                                "🔒 Override", width="small"),
                        },
                    )

                    if st.button(
                        "💾 Simpan Semua Perubahan Absensi",
                        key=f"_btn_save_{account}_{periode}",
                        type="primary",
                    ):
                        _saved_n, _err_list = 0, []
                        for _, _row in _edited_df.iterrows():
                            try:
                                update_absensi_row(
                                    account            = account,
                                    tanggal            = _row["Tanggal"],
                                    jam_masuk          = str(_row["Jam Masuk"]   or ""),
                                    jam_keluar         = str(_row["Jam Keluar"]  or ""),
                                    status_klasifikasi = str(_row["Klasifikasi"] or "None"),
                                    catatan            = str(_row["Catatan"]     or ""),
                                    is_manual_override = 1 if _row["Override"] else 0,
                                )
                                _saved_n += 1
                            except Exception as _e_row:
                                _err_list.append(f"{_row['Tanggal']}: {_e_row}")
                        st.cache_data.clear()
                        if _err_list:
                            st.warning(
                                f"⚠️ {_saved_n} baris tersimpan, "
                                f"{len(_err_list)} gagal. Error pertama: {_err_list[0]}"
                            )
                        else:
                            st.success(
                                f"✅ {_saved_n} baris absensi berhasil disimpan ke database."
                            )

    st.markdown("<div style='margin-top:1.5rem'></div>", unsafe_allow_html=True)
    if st.button("❌ Tutup Rincian", use_container_width=True, type="secondary", key=f"btn_close_dlg_{account}"):
        st.session_state.dialog_target = "closed"
        st.rerun()


@st.cache_data(show_spinner=False)
def process_file(file_bytes):
    buf = io.BytesIO(file_bytes)
    df = pd.read_excel(
        buf,
        sheet_name=_get_sheet_name(file_bytes),
        header=3,
        dtype={"Earliest": str, "Latest": str},
    ).rename(columns={"Unnamed: 0": "Time_Date", "Unnamed: 1": "Name", "Unnamed: 2": "Account"})

    required = ["Name", "Account", "Rules", "Shift", "Earliest", "Latest",
                "Attendance results", "Leave & Overtime Application",
                "Number of absences(Count)"]

    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(f"Kolom tidak ditemukan: {missing}")

    k_sick_col    = _find_ksick_col(df)
    al_col        = _find_al_col(df)
    ul_col        = _find_ul_col(df)
    dur_late_col  = _find_duration_late_col(df)
    dur_early_col = _find_duration_early_col(df)
    wfh_col       = _find_wfh_col(df)
    offsite_col   = _find_offsite_col(df)
    missed_punch_col = _find_missed_punch_col(df)

    hl_col           = _find_hl_col(df)
    ml_col           = _find_ml_col(df)
    wml_col          = _find_wml_col(df)
    ot_col           = _find_ot_col(df)
    rl_col           = _find_rl_col(df)
    pl_col           = _find_pl_col(df)

    df = df.copy()
    df = df[df["Account"].notna() & df["Rules"].notna()]
    df = df[~df["Account"].astype(str).str.strip().isin(["", "--"])]

    df["_statuses"] = df.apply(
        lambda r: classify(
            r["Earliest"], r["Shift"], r["Attendance results"],
            latest_raw=r["Latest"],
            leave_app=r["Leave & Overtime Application"],
            absences_count=r["Number of absences(Count)"],
            k_sick_count=r.get(k_sick_col)     if k_sick_col    else None,
            al_count=r.get(al_col)              if al_col        else None,
            ul_count=r.get(ul_col)              if ul_col        else None,
            duration_late=r.get(dur_late_col)   if dur_late_col  else None,
            duration_early=r.get(dur_early_col) if dur_early_col else None,
            wfh_count=r.get(wfh_col)            if wfh_col       else None,
            offsite_hour=r.get(offsite_col)     if offsite_col   else None,
            missed_punch_count=r.get(missed_punch_col) if missed_punch_col else None,
            hl_count=r.get(hl_col)   if hl_col   else None,
            ml_count=r.get(ml_col)   if ml_col   else None,
            wml_count=r.get(wml_col) if wml_col  else None,
            ot_count=r.get(ot_col)   if ot_col   else None,
            rl_count=r.get(rl_col)   if rl_col   else None,
        ),
        axis=1,
    )

    all_employees = (
        df.groupby("Account")["Rules"]
        .agg(lambda x: x.mode()[0])
        .reset_index()
    )

    df_classified = df[df["_statuses"].notna()].copy()
    df_exploded   = df_classified.explode("_statuses").rename(columns={"_statuses": "Status"})

    pivot = df_exploded.pivot_table(
        index="Account",
        columns="Status",
        values="Shift",
        aggfunc="count",
        fill_value=0,
    ).reset_index()
    pivot.columns.name = None

    ALL_STATUS_COLS = ["S", "Late", "1/2 UL", "UL", "AL", "1/2 AL", "WFA", "1/2 WFA", "WFS", "DW", "K", "Off",
                       "HL", "ML", "WML", "OT", "RL", "H"]
    for col in ALL_STATUS_COLS:
        if col not in pivot.columns:
            pivot[col] = 0

    pivot = all_employees.merge(pivot, on="Account", how="left").fillna(0)
    for col in ALL_STATUS_COLS:
        pivot[col] = pivot[col].astype(int)

    name_map = df.groupby("Account")["Name"].first()
    pivot["Nama"] = pivot["Account"].map(name_map)

    pivot = pivot.sort_values(["Rules", "Nama"]).reset_index(drop=True)
    pivot.insert(0, "No.", range(1, len(pivot) + 1))
    result = pivot[["No.", "Nama", "Account", "Rules",
                    "S", "Late", "1/2 UL", "UL", "AL", "1/2 AL",
                    "WFA", "1/2 WFA", "WFS",
                    "DW", "K", "Off",
                    "HL", "ML", "WML", "OT", "RL"]].copy()

    stats = {
        "total_rows": len(df),
        "classified": len(df_classified),
        "skipped"   : len(df) - len(df_classified),
        "employees" : len(result),
        "dist"      : df_exploded["Status"].value_counts(dropna=False).to_dict(),
    }
    return result, stats


# ──────────────────────────────────────────────────────────────
# Ekspor Kalender Harian (.xlsx)
# Format sesuai sampleexpor.xlsx:
#   Baris 1 : NO | KTP | NAME | tgl-pertama (format 'd') | =D1+1 | ...
#   Baris 2 : (kosong x3) | =D1 (format 'ddd') | =E1 | ...
#   Baris 3+: data karyawan
# Setiap sel diberi background color sesuai klasifikasi (palet pastel).
# ──────────────────────────────────────────────────────────────

def _populate_calendar_ws(ws, df_daily, df_employees):
    """
    Tulis data kalender harian ke sebuah openpyxl Worksheet yang sudah ada.
    Dipakai bersama oleh to_excel_calendar_bytes() dan export_multi_period_bytes().
    """
    thin   = Side(style="thin", color="000000")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    CENTER = Alignment(horizontal="center", vertical="center")
    BOLD   = Font(name="Arial", bold=True, size=10)
    PLAIN  = Font(name="Arial", bold=False, size=9)

    dates = sorted(df_daily["Date"].dropna().unique()) if not df_daily.empty else []

    daily_map: dict = {}
    for _, row in df_daily.iterrows():
        acc = row["Account"]
        if acc not in daily_map:
            daily_map[acc] = {}
        daily_map[acc][row["Date"]] = (
            row.get("Shift", ""),
            row.get("Classification"),
            str(row.get("Remarks", "") or "").strip(),
        )

    n_date_cols = len(dates)

    # ── Baris 1: header kolom + tanggal (format 'd') ───────────────────
    for ci, header in enumerate(["NO", "KTP", "NAME", "ACCOUNT"], 1):
        c = ws.cell(1, ci)
        c.value     = header
        c.font      = BOLD
        c.alignment = CENTER
        c.border    = BORDER

    if dates:
        try:
            first_dt = _dt.date.fromisoformat(dates[0])
            c = ws.cell(1, 5)
            c.value         = _dt.datetime(first_dt.year, first_dt.month, first_dt.day)
            c.number_format = 'd'
            c.font          = BOLD
            c.alignment     = CENTER
            c.border        = BORDER
        except Exception:
            pass

        for di in range(1, n_date_cols):
            prev_col = get_column_letter(5 + di - 1)
            c = ws.cell(1, 5 + di)
            c.value         = f"={prev_col}1+1"
            c.number_format = 'd'
            c.font          = BOLD
            c.alignment     = CENTER
            c.border        = BORDER

    # ── Baris 2: singkatan hari (format 'ddd') ─────────────────────────
    for ci in range(1, 5):
        c = ws.cell(2, ci)
        c.font      = BOLD
        c.alignment = CENTER
        c.border    = BORDER

    for di in range(n_date_cols):
        date_col = get_column_letter(5 + di)
        c = ws.cell(2, 5 + di)
        c.value         = f"={date_col}1"
        c.number_format = 'ddd'
        c.font          = BOLD
        c.alignment     = CENTER
        c.border        = BORDER

    # ── Baris 3+: data per karyawan ────────────────────────────────────
    emp_list = df_employees[["Nama", "Account"]].drop_duplicates("Account").to_dict("records")

    for ri, emp in enumerate(emp_list):
        er  = ri + 3
        acc = emp["Account"]

        for ci_fix, val_fix in [(1, ri + 1), (2, None), (3, emp["Nama"]), (4, acc)]:
            c = ws.cell(er, ci_fix)
            c.value     = val_fix
            c.font      = PLAIN
            c.alignment = CENTER
            c.border    = BORDER

        for di, d in enumerate(dates):
            ci = 5 + di
            c  = ws.cell(er, ci)
            c.border    = BORDER
            c.alignment = CENTER

            _cell_data = daily_map.get(acc, {}).get(d, ("", None, ""))
            shift_t = _cell_data[0]
            klas    = _cell_data[1]
            _rmk    = _cell_data[2] if len(_cell_data) > 2 else ""
            label = _get_cell_display(shift_t, klas)

            c.value = label
            c.font  = Font(name="Arial", size=9, bold=(label in ("DW", "K")))
            if label and label in _CELL_FILL:
                c.fill = _CELL_FILL[label]

            # ── Remarks → openpyxl Comment ──────────────────────────
            if _rmk:
                _comment = Comment(text=_rmk, author="Absensi Rekap")
                _comment.width  = 200
                _comment.height = 80
                c.comment = _comment

    # ── Lebar kolom ────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 13.0
    ws.column_dimensions["B"].width = 25.7
    ws.column_dimensions["C"].width = 28.1
    ws.column_dimensions["D"].width = 22.0
    for di in range(n_date_cols):
        ws.column_dimensions[get_column_letter(5 + di)].width = 13.0


def to_excel_calendar_bytes(df_daily, df_employees, time_range=""):
    """Buat file .xlsx kalender harian (satu sheet) — hasil upload / filter."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Absensi"
    _populate_calendar_ws(ws, df_daily, df_employees)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@st.cache_data(show_spinner=False)
def export_multi_period_bytes(selected_periods: tuple) -> bytes:
    """
    Buat satu file .xlsx dengan beberapa sheet — satu sheet per periode yang dipilih.
    Data diambil dari database (tidak memerlukan file Excel di-upload ulang).
    """

    wb = Workbook()
    wb.remove(wb.active)   # hapus sheet kosong default

    for periode in selected_periods:
        df_daily = _get_all_daily_from_db(periode)
        df_rekap_raw = get_rekap(periode)
        if df_rekap_raw.empty:
            continue

        df_emp = df_rekap_raw.rename(columns={
            "nama": "Nama", "account": "Account", "rules": "Rules",
        })
        # Pastikan kolom Nama & Account tersedia
        for _col in ("Nama", "Account"):
            if _col not in df_emp.columns:
                df_emp[_col] = ""

        # Sheet name maks 31 karakter (batasan Excel)
        sheet_title = str(periode)[:31]
        # Hindari duplikat nama sheet
        existing_titles = [s.title for s in wb.worksheets]
        if sheet_title in existing_titles:
            sheet_title = sheet_title[:27] + f"_{len(existing_titles)}"

        ws = wb.create_sheet(title=sheet_title)
        _populate_calendar_ws(ws, df_daily, df_emp)

    if not wb.worksheets:
        ws = wb.create_sheet("Tidak Ada Data")
        ws.cell(1, 1).value = "Tidak ada data tersedia untuk periode yang dipilih."

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ──────────────────────────────────────────────────────────────
# Definisi kolom tampilan tabel
# ──────────────────────────────────────────────────────────────

CORE_COLS = [
    "No.", "Nama", "Account", "Rules",
    "S", "Late", "1/2 UL", "UL", "DW",
    "K", "AL", "1/2 AL", "WFA", "1/2 WFA", "WFS", "Off",
    "HL", "ML", "WML", "OT", "RL", "H", "PL",
]

OPTIONAL_COLS_DEF = [
    ("K",      "💊 K (Sakit)",   "Sakit dgn Surat"),
    ("AL",     "🌴 AL",          "Annual Leave penuh"),
    ("1/2 AL", "🌗 1/2 AL",      "Annual Leave setengah hari"),
    ("WFA",    "🏠 WFA",         "Work From Home penuh"),
    ("1/2 WFA","🏡 1/2 WFA",     "Work From Home setengah hari"),
    ("WFS",    "📍 WFS",         "Work From Offsite"),
    ("Off",    "🏖️ Off",         "Rest / Not scheduled"),
    ("HL",     "💍 HL",           "Cuti Pernikahan"),
    ("ML",     "🤱 ML",           "Cuti Melahirkan"),
    ("WML",    "👶 WML",          "Cuti Istri Melahirkan"),
    ("OT",     "📝 OT",           "Cuti Lainnya"),
    ("RL",     "📅 RL",           "Roster Leave"),
    ("H",      "🔴 H",            "Hari Libur Nasional"),
    ("PL",     "🪪 PL",           "Personal Leave TKA"),
]
OPTIONAL_KEYS   = [c[0] for c in OPTIONAL_COLS_DEF]
OPTIONAL_LABELS = {c[0]: c[1] for c in OPTIONAL_COLS_DEF}
OPTIONAL_DESCS  = {c[0]: c[2] for c in OPTIONAL_COLS_DEF}

COL_CONFIG_ALL = {
    "No."     : st.column_config.NumberColumn("No.", width="small"),
    "Nama"    : st.column_config.TextColumn("Nama", width="large"),
    "Account" : st.column_config.TextColumn("Account", width="medium"),
    "Rules"   : st.column_config.TextColumn("Rules", width="medium"),
    "S"       : st.column_config.NumberColumn("📋 S (Shift)",  format="%d", width="small"),
    "Late"    : st.column_config.NumberColumn("🕐 Late",       format="%d", width="small"),
    "1/2 UL"  : st.column_config.NumberColumn("⛔ 1/2 UL",    format="%d", width="small"),
    "UL"      : st.column_config.NumberColumn("📋 UL",         format="%d", width="small"),
    "DW"      : st.column_config.NumberColumn("🚫 DW",         format="%d", width="small"),
    "K"       : st.column_config.NumberColumn("💊 K",          format="%d", width="small"),
    "AL"      : st.column_config.NumberColumn("🌴 AL",         format="%d", width="small"),
    "1/2 AL"  : st.column_config.NumberColumn("🌗 1/2 AL",     format="%d", width="small"),
    "WFA"     : st.column_config.NumberColumn("🏠 WFA",        format="%d", width="small"),
    "1/2 WFA" : st.column_config.NumberColumn("🏡 1/2 WFA",    format="%d", width="small"),
    "WFS"     : st.column_config.NumberColumn("📍 WFS",        format="%d", width="small"),
    "Off"     : st.column_config.NumberColumn("🏖️ Off",        format="%d", width="small"),
    "HL"      : st.column_config.NumberColumn("💍 HL",          format="%d", width="small"),
    "ML"      : st.column_config.NumberColumn("🤱 ML",          format="%d", width="small"),
    "WML"     : st.column_config.NumberColumn("👶 WML",         format="%d", width="small"),
    "OT"      : st.column_config.NumberColumn("📝 OT",          format="%d", width="small"),
    "RL"      : st.column_config.NumberColumn("📅 RL",          format="%d", width="small"),
    "H"       : st.column_config.NumberColumn("🔴 H",           format="%d", width="small"),
    "PL"      : st.column_config.NumberColumn("🪪 PL",           format="%d", width="small"),
}



# ──────────────────────────────────────────────────────────────
# Dialog: Logic Klasifikasi
# ──────────────────────────────────────────────────────────────

_LOGIC_HTML = (
    '<div style="font-size:0.85rem;color:#334155;line-height:1.9;">'

    '<div style="background:#f0f9ff;border-radius:8px;padding:0.7rem 1rem;margin-bottom:1.2rem;'
    'font-size:0.82rem;border-left:3px solid #0ea5e9;">'
    '<b>🗄️ Penyimpanan Data (Database)</b><br>'
    'Setiap kali file Excel diupload, <b>semua data detail harian</b> disimpan ke database SQLite secara otomatis. '
    'Data yang tersimpan meliputi: tanggal, shift, tipe shift (Normal/Off), jam masuk, jam keluar, '
    'jam kerja, status absensi (Attendance Results), status klasifikasi, dan data leave. '
    'Periode yang sudah tersimpan dapat dipilih kembali dari dropdown tanpa perlu upload ulang.'
    '</div>'

    '<div style="font-weight:700;color:#0f172a;margin-bottom:0.4rem;font-size:0.82rem;'
    'text-transform:uppercase;letter-spacing:0.06em;">📊 Format Ekspor Kalender Harian (.xlsx)</div>'

    '<div style="background:#f0fdf4;border-radius:8px;padding:0.7rem 1rem;margin-bottom:1.2rem;'
    'font-size:0.82rem;border-left:3px solid #22c55e;">'
    '<b>Layout:</b> Baris = karyawan, Kolom = tanggal (1 s/d akhir bulan)<br>'
    '<b>Struktur baris header:</b><br>'
    '&nbsp;&nbsp;• Baris 1 — <b>NO</b> | <b>KTP</b> (dikosongkan) | <b>NAME</b> | '
    'tanggal pertama (format <code>d</code> → angka hari) | <code>=D1+1</code> | <code>=E1+1</code> | …<br>'
    '&nbsp;&nbsp;• Baris 2 — (kosong) | (kosong) | (kosong) | <code>=D1</code> (format <code>ddd</code> → Tue/Wed…) | …<br>'
    '&nbsp;&nbsp;• Baris 3+ — data harian per karyawan<br><br>'
    '<b>Isi sel harian:</b><br>'
    '&nbsp;&nbsp;• Status S (hadir normal, semua tipe shift) → <b>"S"</b><br>'
    '&nbsp;&nbsp;• Status khusus → kode label (tabel di bawah)<br>'
    '&nbsp;&nbsp;• Tidak ada data / None → kosong (tanpa warna)<br><br>'
    '<b>Styling:</b> Font Arial, semua sel center-aligned, thin border hitam. '
    'Setiap klasifikasi diberi <b>background color pastel</b> (lihat warna di bawah). '
    'Lebar kolom: A=13, B=25.7, C=28.1, kolom tanggal=13.'
    '</div>'

    '<div style="font-weight:700;color:#0f172a;margin-bottom:0.6rem;font-size:0.82rem;'
    'text-transform:uppercase;letter-spacing:0.06em;">🎨 Warna Sel Ekspor</div>'
    
    '<tr><td style="padding:0.3rem 0.7rem;"><b>H</b></td>'
    '<td style="padding:0.3rem 0.7rem;"><span style="background:#FF9999;padding:2px 10px;border-radius:3px;">▮ #FF9999</span></td>'
    '<td style="padding:0.3rem 0.7rem;">Hari Libur Nasional — merah cerah</td></tr>'

    '<table style="width:100%;border-collapse:collapse;margin-bottom:1.2rem;">'
    '<tr style="background:#f1f5f9;">'
    '<td style="padding:0.4rem 0.7rem;font-weight:600;width:80px;">Label</td>'
    '<td style="padding:0.4rem 0.7rem;font-weight:600;width:120px;">Warna Sel</td>'
    '<td style="padding:0.4rem 0.7rem;font-weight:600;">Status Klasifikasi</td>'
    '</tr>'
    '<tr><td style="padding:0.3rem 0.7rem;"><b>S</b></td>'
    '<td style="padding:0.3rem 0.7rem;"><span style="background:#D9EAD3;padding:2px 10px;border-radius:3px;">▮ #D9EAD3</span></td>'
    '<td style="padding:0.3rem 0.7rem;">Hadir normal — hijau muda</td></tr>'
    '<tr style="background:#f8fafc;"><td style="padding:0.3rem 0.7rem;"><b>OFF</b></td>'
    '<td style="padding:0.3rem 0.7rem;"><span style="background:#CFE2F3;padding:2px 10px;border-radius:3px;">▮ #CFE2F3</span></td>'
    '<td style="padding:0.3rem 0.7rem;">Off / libur — biru muda</td></tr>'
    '<tr><td style="padding:0.3rem 0.7rem;"><b>AL</b></td>'
    '<td style="padding:0.3rem 0.7rem;"><span style="background:#D9D2E9;padding:2px 10px;border-radius:3px;">▮ #D9D2E9</span></td>'
    '<td style="padding:0.3rem 0.7rem;">Annual Leave penuh — ungu muda</td></tr>'
    '<tr style="background:#f8fafc;"><td style="padding:0.3rem 0.7rem;"><b>0,5AL</b></td>'
    '<td style="padding:0.3rem 0.7rem;"><span style="background:#EAD1DC;padding:2px 10px;border-radius:3px;">▮ #EAD1DC</span></td>'
    '<td style="padding:0.3rem 0.7rem;">Annual Leave ½ hari — pink</td></tr>'
    '<tr><td style="padding:0.3rem 0.7rem;"><b>WFA</b></td>'
    '<td style="padding:0.3rem 0.7rem;"><span style="background:#D0E4F7;padding:2px 10px;border-radius:3px;">▮ #D0E4F7</span></td>'
    '<td style="padding:0.3rem 0.7rem;">Work From Home penuh — biru langit</td></tr>'
    '<tr style="background:#f8fafc;"><td style="padding:0.3rem 0.7rem;"><b>0,5WFA</b></td>'
    '<td style="padding:0.3rem 0.7rem;"><span style="background:#C9DAF8;padding:2px 10px;border-radius:3px;">▮ #C9DAF8</span></td>'
    '<td style="padding:0.3rem 0.7rem;">Work From Home ½ hari — biru medium</td></tr>'
    '<tr><td style="padding:0.3rem 0.7rem;"><b>WFS</b></td>'
    '<td style="padding:0.3rem 0.7rem;"><span style="background:#B4C7E7;padding:2px 10px;border-radius:3px;">▮ #B4C7E7</span></td>'
    '<td style="padding:0.3rem 0.7rem;">Work From Offsite — biru indigo</td></tr>'
    '<tr style="background:#f8fafc;"><td style="padding:0.3rem 0.7rem;"><b>UL</b></td>'
    '<td style="padding:0.3rem 0.7rem;"><span style="background:#B7E1CD;padding:2px 10px;border-radius:3px;">▮ #B7E1CD</span></td>'
    '<td style="padding:0.3rem 0.7rem;">Unpaid Leave penuh — hijau sedang</td></tr>'
    '<tr><td style="padding:0.3rem 0.7rem;"><b>0,5UL</b></td>'
    '<td style="padding:0.3rem 0.7rem;"><span style="background:#FCE5CD;padding:2px 10px;border-radius:3px;">▮ #FCE5CD</span></td>'
    '<td style="padding:0.3rem 0.7rem;">Unpaid Leave ½ hari / terlambat &gt;120 mnt — oranye muda</td></tr>'
    '<tr style="background:#f8fafc;"><td style="padding:0.3rem 0.7rem;"><b>L</b></td>'
    '<td style="padding:0.3rem 0.7rem;"><span style="background:#FFF2CC;padding:2px 10px;border-radius:3px;">▮ #FFF2CC</span></td>'
    '<td style="padding:0.3rem 0.7rem;">Late / pulang cepat 1–120 mnt — kuning</td></tr>'
    '<tr><td style="padding:0.3rem 0.7rem;"><b>K</b></td>'
    '<td style="padding:0.3rem 0.7rem;"><span style="background:#F4CCCC;padding:2px 10px;border-radius:3px;">▮ #F4CCCC</span></td>'
    '<td style="padding:0.3rem 0.7rem;">Sakit dengan surat — merah muda</td></tr>'
    '<tr style="background:#f8fafc;"><td style="padding:0.3rem 0.7rem;"><b>DW</b></td>'
    '<td style="padding:0.3rem 0.7rem;"><span style="background:#EA9999;padding:2px 10px;border-radius:3px;">▮ #EA9999</span></td>'
    '<td style="padding:0.3rem 0.7rem;">Tidak hadir / Absence — merah</td></tr>'
    '<tr><td style="padding:0.3rem 0.7rem;"><b>HL</b></td>'
    '<td style="padding:0.3rem 0.7rem;"><span style="background:#FFE599;padding:2px 10px;border-radius:3px;">▮ #FFE599</span></td>'
    '<td style="padding:0.3rem 0.7rem;">Cuti Pernikahan — kuning emas</td></tr>'
    '<tr style="background:#f8fafc;"><td style="padding:0.3rem 0.7rem;"><b>ML</b></td>'
    '<td style="padding:0.3rem 0.7rem;"><span style="background:#B6D7A8;padding:2px 10px;border-radius:3px;">▮ #B6D7A8</span></td>'
    '<td style="padding:0.3rem 0.7rem;">Cuti Melahirkan — hijau sedang</td></tr>'
    '<tr><td style="padding:0.3rem 0.7rem;"><b>WML</b></td>'
    '<td style="padding:0.3rem 0.7rem;"><span style="background:#A2C4C9;padding:2px 10px;border-radius:3px;">▮ #A2C4C9</span></td>'
    '<td style="padding:0.3rem 0.7rem;">Cuti Istri Melahirkan — teal</td></tr>'
    '<tr style="background:#f8fafc;"><td style="padding:0.3rem 0.7rem;"><b>OT</b></td>'
    '<td style="padding:0.3rem 0.7rem;"><span style="background:#D9D9D9;padding:2px 10px;border-radius:3px;">▮ #D9D9D9</span></td>'
    '<td style="padding:0.3rem 0.7rem;">Cuti Lainnya — abu-abu</td></tr>'
    '<tr><td style="padding:0.3rem 0.7rem;"><i>(kosong)</i></td>'
    '<td style="padding:0.3rem 0.7rem;"><span style="background:#FFFFFF;border:1px solid #e2e8f0;padding:2px 10px;border-radius:3px;">▮ putih</span></td>'
    '<td style="padding:0.3rem 0.7rem;">None — tidak memenuhi kondisi manapun</td></tr>'
    '</table>'

    '<div style="font-weight:700;color:#0f172a;margin-bottom:0.4rem;font-size:0.82rem;'
    'text-transform:uppercase;letter-spacing:0.06em;">📅 Tipe Shift</div>'

    '<table style="width:100%;border-collapse:collapse;margin-bottom:1.2rem;">'
    '<tr style="background:#f1f5f9;">'
    '<td style="padding:0.4rem 0.7rem;font-weight:600;white-space:nowrap;width:110px;">☀️ Shift</td>'
    '<td style="padding:0.4rem 0.7rem;">Semua shift kerja — termasuk shift pagi, malam, S1, S2, Night, dll.</td>'
    '</tr>'
    '<tr>'
    '<td style="padding:0.4rem 0.7rem;font-weight:600;white-space:nowrap;">🏖️ Off</td>'
    '<td style="padding:0.4rem 0.7rem;">Shift <code>"Rest"</code> — hari libur atau tidak terjadwal.</td>'
    '</tr>'
    '</table>'

    '<div style="font-weight:700;color:#0f172a;margin-bottom:0.4rem;font-size:0.82rem;'
    'text-transform:uppercase;letter-spacing:0.06em;">📂 Kolom Sumber Data Klasifikasi</div>'

    '<table style="width:100%;border-collapse:collapse;margin-bottom:1.2rem;">'
    '<tr style="background:#f1f5f9;">'
    '<td style="padding:0.4rem 0.7rem;font-weight:600;white-space:nowrap;width:130px;">Status</td>'
    '<td style="padding:0.4rem 0.7rem;font-weight:600;">Kolom Sumber</td>'
    '<td style="padding:0.4rem 0.7rem;font-weight:600;">Nilai Pemicu</td>'
    '</tr>'
    '<tr>'
    '<td style="padding:0.4rem 0.7rem;">📍 WFS</td>'
    '<td style="padding:0.4rem 0.7rem;"><code>Attendance results</code> + <code>Offsite(Hour)</code></td>'
    '<td style="padding:0.4rem 0.7rem;">att = <b>"Normal（Offsite）"</b> atau <b>"Normal（Correction of missed punch、Offsite）"</b> DAN Offsite(Hour) &ne; "--"/kosong</td>'
    '</tr>'
    '<tr style="background:#f1f5f9;">'
    '<td style="padding:0.4rem 0.7rem;">🏠 WFA</td>'
    '<td style="padding:0.4rem 0.7rem;"><code>WFH-WorkFromHome-家办公(Day(s))</code></td>'
    '<td style="padding:0.4rem 0.7rem;">nilai = <b>1</b></td>'
    '</tr>'
    '<tr>'
    '<td style="padding:0.4rem 0.7rem;">🏡 1/2 WFA</td>'
    '<td style="padding:0.4rem 0.7rem;"><code>WFH-WorkFromHome-家办公(Day(s))</code></td>'
    '<td style="padding:0.4rem 0.7rem;">nilai = <b>0.5</b></td>'
    '</tr>'
    '<tr style="background:#f1f5f9;">'
    '<td style="padding:0.4rem 0.7rem;">🌴 AL</td>'
    '<td style="padding:0.4rem 0.7rem;"><code>AnnualLeave - 印尼员工年假(Day(s))</code></td>'
    '<td style="padding:0.4rem 0.7rem;">nilai = <b>1</b></td>'
    '</tr>'
    '<tr>'
    '<td style="padding:0.4rem 0.7rem;">🌗 1/2 AL</td>'
    '<td style="padding:0.4rem 0.7rem;"><code>AnnualLeave - 印尼员工年假(Day(s))</code></td>'
    '<td style="padding:0.4rem 0.7rem;">nilai = <b>0.5</b></td>'
    '</tr>'
    '<tr style="background:#f1f5f9;">'
    '<td style="padding:0.4rem 0.7rem;">📋 UL</td>'
    '<td style="padding:0.4rem 0.7rem;"><code>UL-Unpaid Leave-事假(Day(s))</code></td>'
    '<td style="padding:0.4rem 0.7rem;">nilai = <b>1</b></td>'
    '</tr>'
    '<tr>'
    '<td style="padding:0.4rem 0.7rem;">📋 1/2 UL (dari kolom)</td>'
    '<td style="padding:0.4rem 0.7rem;"><code>UL-Unpaid Leave-事假(Day(s))</code></td>'
    '<td style="padding:0.4rem 0.7rem;">nilai = <b>0.5</b></td>'
    '</tr>'
    '<tr style="background:#f1f5f9;">'
    '<td style="padding:0.4rem 0.7rem;">🕐 Late (masuk)</td>'
    '<td style="padding:0.4rem 0.7rem;"><code>Duration of late arrival(分钟)</code></td>'
    '<td style="padding:0.4rem 0.7rem;">1 – 120 menit</td>'
    '</tr>'
    '<tr>'
    '<td style="padding:0.4rem 0.7rem;">⛔ 1/2 UL (masuk)</td>'
    '<td style="padding:0.4rem 0.7rem;"><code>Duration of late arrival(分钟)</code></td>'
    '<td style="padding:0.4rem 0.7rem;">&gt; 120 menit</td>'
    '</tr>'
    '<tr style="background:#f1f5f9;">'
    '<td style="padding:0.4rem 0.7rem;">🕐 Late (pulang)</td>'
    '<td style="padding:0.4rem 0.7rem;"><code>Duration of early departure(分钟)</code></td>'
    '<td style="padding:0.4rem 0.7rem;">1 – 120 menit</td>'
    '</tr>'
    '<tr>'
    '<td style="padding:0.4rem 0.7rem;">⛔ 1/2 UL (pulang)</td>'
    '<td style="padding:0.4rem 0.7rem;"><code>Duration of early departure(分钟)</code></td>'
    '<td style="padding:0.4rem 0.7rem;">&gt; 120 menit</td>'
    '</tr>'
    '<tr style="background:#f1f5f9;">'
    '<td style="padding:0.4rem 0.7rem;">💊 K</td>'
    '<td style="padding:0.4rem 0.7rem;"><code>K-Sick W Letter-病假有信(Day(s))</code></td>'
    '<td style="padding:0.4rem 0.7rem;">&ne; 0 / "--"</td>'
    '</tr>'
    '<tr>'
    '<td style="padding:0.4rem 0.7rem;">🚫 DW</td>'
    '<td style="padding:0.4rem 0.7rem;"><code>Number of absences(Count)</code></td>'
    '<td style="padding:0.4rem 0.7rem;">&ne; 0 / "--"</td>'
    '</tr>'
    '<tr style="background:#f1f5f9;">'
    '<td style="padding:0.4rem 0.7rem;">⛔ 1/2 UL (missed punch)</td>'
    '<td style="padding:0.4rem 0.7rem;"><code>Number of missed punches(Count)</code></td>'
    '<td style="padding:0.4rem 0.7rem;">= <b>1</b> (tepat satu punch terlewat)</td>'
    '</tr>'
    '<tr><td style="padding:0.4rem 0.7rem;">💍 HL</td>'
    '<td style="padding:0.4rem 0.7rem;"><code>HL-Happy(Marry)-婚假(Day(s))</code></td>'
    '<td style="padding:0.4rem 0.7rem;">nilai = <b>1</b></td>'
    '</tr>'
    '<tr style="background:#f1f5f9;">'
    '<td style="padding:0.4rem 0.7rem;">🤱 ML</td>'
    '<td style="padding:0.4rem 0.7rem;"><code>ML-MaternityLeave-产假(Day(s))</code></td>'
    '<td style="padding:0.4rem 0.7rem;">&ne; "--" / kosong / 0 (nilai hari melahirkan)</td>'
    '</tr>'
    '<tr>'
    '<td style="padding:0.4rem 0.7rem;">👶 WML</td>'
    '<td style="padding:0.4rem 0.7rem;"><code>WML-WifeMater-妻产假(Day(s))</code></td>'
    '<td style="padding:0.4rem 0.7rem;">nilai = <b>1</b></td>'
    '</tr>'
    '<tr style="background:#f1f5f9;">'
    '<td style="padding:0.4rem 0.7rem;">📝 OT</td>'
    '<td style="padding:0.4rem 0.7rem;"><code>OT - Others - 其他(Day(s))</code></td>'
    '<td style="padding:0.4rem 0.7rem;">nilai = <b>1</b></td>'
    '</tr>'
    '<tr style="background:#f1f5f9;">'
    '<td style="padding:0.4rem 0.7rem;">📅 RL</td>'
    '<td style="padding:0.4rem 0.7rem;"><code>RL - Roster Leave(Day(s))</code></td>'
    '<td style="padding:0.4rem 0.7rem;">&ne; "--" / kosong / 0</td>'
    '</tr>'
    '</table>'

    '<div style="font-weight:700;color:#0f172a;margin-bottom:0.4rem;font-size:0.82rem;'
    'text-transform:uppercase;letter-spacing:0.06em;">📊 Status &amp; Kondisi Pemicu</div>'

    '<table style="width:100%;border-collapse:collapse;margin-bottom:1.2rem;">'

    '<tr style="background:#f1f5f9;">'
    '<td style="padding:0.4rem 0.7rem;font-weight:600;white-space:nowrap;width:110px;">🏖️ Off</td>'
    '<td style="padding:0.4rem 0.7rem;">Att Results bernilai tepat '
    '<code>"Normal (rest)"</code> atau <code>"Normal (not scheduled)"</code> '
    '— dicek <b>paling awal</b>.</td>'
    '</tr>'

    '<tr>'
    '<td style="padding:0.4rem 0.7rem;font-weight:600;white-space:nowrap;">📍 WFS</td>'
    '<td style="padding:0.4rem 0.7rem;">'
    'Att Results = <b>tepat</b> <code>"Normal（Offsite）"</code> atau '
    '<code>"Normal（Correction of missed punch、Offsite）"</code> <b>DAN</b> '
    'kolom <code>Offsite(Hour)</code> berisi nilai apapun selain <code>"--"</code> / kosong. '
    'Catatan: menggunakan tanda kurung full-width <code>（）</code> khas output Excel, '
    'bukan ASCII <code>()</code>. '
    'Dicek <b>sebelum</b> skip-shift agar tidak terlewat meski shift kosong.</td>'
    '</tr>'

    '<tr style="background:#f1f5f9;">'
    '<td style="padding:0.4rem 0.7rem;font-weight:600;white-space:nowrap;">💊 K</td>'
    '<td style="padding:0.4rem 0.7rem;">Kolom <code>"K-Sick W Letter"</code> '
    'bernilai <b>bukan</b> <code>"0"</code> atau <code>"--"</code> — sakit dengan surat.<br>'
    'Dicek <em>sebelum</em> DW agar sakit-dengan-surat tidak tertimpa DW.</td>'
    '</tr>'

    '<tr>'
    '<td style="padding:0.4rem 0.7rem;font-weight:600;white-space:nowrap;">🚫 DW</td>'
    '<td style="padding:0.4rem 0.7rem;">Kolom <code>"Number of absences(Count)"</code> '
    'bernilai <b>bukan</b> <code>"0"</code> atau <code>"--"</code> — karyawan tidak hadir.</td>'
    '</tr>'

    '<tr style="background:#f1f5f9;">'
    '<td style="padding:0.4rem 0.7rem;font-weight:600;white-space:nowrap;">🌴 AL</td>'
    '<td style="padding:0.4rem 0.7rem;">Kolom <code>AnnualLeave</code> = <b>1</b> → <code>["AL"]</code></td>'
    '</tr>'

    '<tr>'
    '<td style="padding:0.4rem 0.7rem;font-weight:600;white-space:nowrap;">🌗 1/2 AL</td>'
    '<td style="padding:0.4rem 0.7rem;">Kolom <code>AnnualLeave</code> = <b>0.5</b> → <code>["1/2 AL"]</code></td>'
    '</tr>'

    '<tr style="background:#f1f5f9;">'
    '<td style="padding:0.4rem 0.7rem;font-weight:600;white-space:nowrap;">📋 UL</td>'
    '<td style="padding:0.4rem 0.7rem;">Kolom <code>UL-Unpaid Leave</code> = <b>1</b> — cuti tidak dibayar satu hari penuh.</td>'
    '</tr>'

    '<tr>'
    '<td style="padding:0.4rem 0.7rem;font-weight:600;white-space:nowrap;">⛔ 1/2 UL (dari UL)</td>'
    '<td style="padding:0.4rem 0.7rem;">Kolom <code>UL-Unpaid Leave</code> = <b>0.5</b> — Unpaid Leave setengah hari.</td>'
    '</tr>'

    '<tr style="background:#f1f5f9;">'
    '<td style="padding:0.4rem 0.7rem;font-weight:600;white-space:nowrap;">🏠 WFA</td>'
    '<td style="padding:0.4rem 0.7rem;">Kolom <code>WFH-WorkFromHome</code> = <b>1</b> → Work From Home penuh.</td>'
    '</tr>'

    '<tr>'
    '<td style="padding:0.4rem 0.7rem;font-weight:600;white-space:nowrap;">🏡 1/2 WFA</td>'
    '<td style="padding:0.4rem 0.7rem;">Kolom <code>WFH-WorkFromHome</code> = <b>0.5</b> → Work From Home setengah hari.</td>'
    '</tr>'

    '<tr>'
    '<td style="padding:0.4rem 0.7rem;font-weight:600;white-space:nowrap;">💍 HL</td>'
    '<td style="padding:0.4rem 0.7rem;">Kolom <code>HL-Happy(Marry)</code> = <b>1</b> — cuti pernikahan satu hari.</td>'
    '</tr>'
    '<tr style="background:#f1f5f9;">'
    '<td style="padding:0.4rem 0.7rem;font-weight:600;white-space:nowrap;">🤱 ML</td>'
    '<td style="padding:0.4rem 0.7rem;">Kolom <code>ML-MaternityLeave</code> berisi nilai <b>selain "--" / kosong / 0</b> '
    '— jumlah hari cuti melahirkan berapapun memicu ML. Dicek setelah HL.</td>'
    '</tr>'
    '<tr>'
    '<td style="padding:0.4rem 0.7rem;font-weight:600;white-space:nowrap;">👶 WML</td>'
    '<td style="padding:0.4rem 0.7rem;">Kolom <code>WML-WifeMater</code> = <b>1</b> — cuti suami saat istri melahirkan.</td>'
    '</tr>'
    '<tr style="background:#f1f5f9;">'
    '<td style="padding:0.4rem 0.7rem;font-weight:600;white-space:nowrap;">📝 OT</td>'
    '<td style="padding:0.4rem 0.7rem;">Kolom <code>OT - Others</code> = <b>1</b> — jenis cuti lainnya yang tidak masuk kategori di atas.</td>'
    '</tr>'

    '<tr style="background:#f1f5f9;">'
    '<td style="padding:0.4rem 0.7rem;font-weight:600;white-space:nowrap;">🕐 Late</td>'
    '<td style="padding:0.4rem 0.7rem;">'
    'Kedua kolom <code>Duration of late arrival</code> dan <code>Duration of early departure</code> '
    'dievaluasi. Jika keduanya bernilai <b>1–120 menit</b> (tidak ada yang &gt;120 mnt) → <b>Late</b>.</td>'
    '</tr>'

    '<tr>'
    '<td style="padding:0.4rem 0.7rem;font-weight:600;white-space:nowrap;">⛔ 1/2 UL (durasi)</td>'
    '<td style="padding:0.4rem 0.7rem;">'
    'Jika <b>salah satu</b> dari <code>Duration of late arrival</code> atau '
    '<code>Duration of early departure</code> bernilai <b>&gt;120 menit</b>, '
    'maka hasil adalah <b>1/2 UL</b> — meskipun kolom lainnya hanya <b>Late</b>. '
    'Prinsip: ambil yang paling berat.</td>'
    '</tr>'

    '<tr style="background:#f1f5f9;">'
    '<td style="padding:0.4rem 0.7rem;font-weight:600;white-space:nowrap;">⛔ 1/2 UL (missed punch)</td>'
    '<td style="padding:0.4rem 0.7rem;">'
    'Kolom <code>Number of missed punches(Count)</code> tepat bernilai <b>1</b> — '
    'satu punch (masuk atau keluar) tidak terekam. Dicek setelah cek durasi '
    'agar tidak tumpang tindih dengan Late/½UL berbasis durasi.</td>'
    '</tr>'

    '<tr>'
    '<td style="padding:0.4rem 0.7rem;font-weight:600;white-space:nowrap;">📋 S (Shift)</td>'
    '<td style="padding:0.4rem 0.7rem;">Att Results bernilai <b>TEPAT</b> <code>"Normal"</code> '
    'atau <code>"Normal（Correction of missed punch）"</code>.</td>'
    '</tr>'

    '<tr>'
    '<td style="padding:0.4rem 0.7rem;font-weight:600;white-space:nowrap;">❓ None</td>'
    '<td style="padding:0.4rem 0.7rem;">Tidak memenuhi satu pun kondisi di atas — sel ekspor <b>kosong</b> (tanpa warna).</td>'
    '</tr>'

    '</table>'

    '<div style="font-weight:700;color:#0f172a;margin-bottom:0.4rem;font-size:0.82rem;'
    'text-transform:uppercase;letter-spacing:0.06em;">🔀 Alur Keputusan (Urutan Prioritas)</div>'

    '<div style="background:#fff;border:1px solid #e2e8f0;border-radius:8px;'
    'padding:0.8rem 1rem;font-family:monospace;font-size:0.78rem;line-height:2.1;'
    'margin-bottom:1.2rem;color:#475569;">'
    '1.  🏖️ Att = "Normal (rest)" / "Normal (not scheduled)" &rarr; <b>Off</b> &mdash; selesai<br>'
    '2.  📍 Att = <code>"Normal（Offsite）"</code> atau <code>"Normal（Correction of missed punch、Offsite）"</code> DAN Offsite(Hour) &ne; "--"/kosong &rarr; <b>WFS</b> &mdash; selesai<br>'
    '3.  ⏭️ Shift = Rest / Not scheduled / kosong / "--" &rarr; <b>dilewati engine</b>, tampil sebagai <b>❓ None</b><br>'
    '4.  💊 Kolom K-Sick W Letter &ne; "0" dan "--" &rarr; <b>K</b> &mdash; selesai<br>'
    '5.  🚫 Kolom Number of absences(Count) &ne; "0" dan "--" &rarr; <b>DW</b> &mdash; selesai<br>'
    '6.  🌴 Kolom AnnualLeave = 0.5 &rarr; <b>1/2 AL</b> &mdash; selesai<br>'
    '&nbsp;&nbsp;&nbsp;Kolom AnnualLeave = 1   &rarr; <b>AL</b> &mdash; selesai<br>'
    '7.  📋 Kolom UL-Unpaid Leave = 1   &rarr; <b>UL</b> &mdash; selesai<br>'
    '&nbsp;&nbsp;&nbsp;Kolom UL-Unpaid Leave = 0.5 &rarr; <b>1/2 UL</b> &mdash; selesai<br>'
    '8.  🏠 Kolom WFH-WorkFromHome = 1   &rarr; <b>WFA</b> &mdash; selesai<br>'
    '&nbsp;&nbsp;&nbsp;Kolom WFH-WorkFromHome = 0.5 &rarr; <b>1/2 WFA</b> &mdash; selesai<br>'
    '9.  💍 Kolom HL-Happy(Marry) = <b>1</b> &rarr; <b>HL</b> &mdash; selesai<br>'
    '10. 🤱 Kolom ML-MaternityLeave &ne; "--"/kosong/0 &rarr; <b>ML</b> &mdash; selesai<br>'
    '11. 👶 Kolom WML-WifeMater = <b>1</b> &rarr; <b>WML</b> &mdash; selesai<br>'
    '12. 📝 Kolom OT - Others = <b>1</b> &rarr; <b>OT</b> &mdash; selesai<br>'
    '13. 📅 Kolom RL-RosterLeave &ne; "--"/kosong/0 &rarr; <b>RL</b> &mdash; selesai<br>'
    '14. 🕐 Kolom Duration of late arrival <b>+</b> Duration of early departure<br>'
    '&nbsp;&nbsp;&nbsp;(Keduanya dievaluasi — diambil yang paling berat)<br>'
    '&nbsp;&nbsp;&nbsp;+-- Salah satu &gt;120 mnt &rarr; <b>1/2 UL</b> &mdash; selesai<br>'
    '&nbsp;&nbsp;&nbsp;+-- Keduanya 1-120 mnt &nbsp;&rarr; <b>Late</b> &mdash; selesai<br>'
    '14.5 💼 Kolom Number of missed punches(Count) = <b>1</b> &rarr; <b>1/2 UL</b> &mdash; selesai<br>'
    '15. 📋 Att TEPAT "Normal" atau "Normal（Correction of missed punch）" &rarr; <b>S</b><br>'
    '16. ❓ Selain itu &rarr; <b>None</b> (sel ekspor kosong, tanpa warna)'
    '</div>'

    '<div style="font-weight:700;color:#0f172a;margin-bottom:0.4rem;font-size:0.82rem;'
    'text-transform:uppercase;letter-spacing:0.06em;">🔄 Perubahan Logika Utama (Update Terbaru)</div>'

    '<div style="background:#fef9ec;border-radius:8px;padding:0.7rem 1rem;margin-bottom:1.2rem;'
    'font-size:0.82rem;border-left:3px solid #eab308;">'

    '<b>1. 🎨 Background Color Sel Ekspor (Update Terbaru):</b><br>'
    '- Setiap sel data di kalender harian kini diberi <b>background color pastel</b> sesuai klasifikasi<br>'
    '- Warna dipilih dari palet Google Sheets yang familiar dan mudah dibedakan<br>'
    '- Sel <b>None</b> (tidak ada data) tetap putih/tanpa warna<br>'
    '- Sel <b>DW</b> dan <b>K</b> menggunakan font bold untuk penekanan visual<br>'
    '- Lihat tabel warna di atas untuk daftar lengkap mapping warna<br><br>'

    '<b>2. 📊 Format Ekspor Kalender Diperbarui (sesuai sampleexpor.xlsx):</b><br>'
    '- <b>Baris 1:</b> NO | KTP | NAME | tanggal pertama (datetime, format <code>d</code>) | <code>=D1+1</code> | <code>=E1+1</code> | …<br>'
    '- <b>Baris 2:</b> (kosong x3) | <code>=D1</code> (format <code>ddd</code>) | <code>=E1</code> | … (singkatan hari otomatis)<br>'
    '- <b>Data mulai baris 3</b> (sebelumnya baris 4 karena ada judul di baris 1)<br>'
    '- <b>Tidak ada baris judul</b> — judul "Rekap Absensi Harian" dihapus<br>'
    '- Kolom KTP dikosongkan (diisi manual jika diperlukan)<br><br>'

    '<b>3. ✏️ Label Sel S Diubah:</b><br>'
    '- <b>Sebelumnya:</b> nama pendek shift (<code>S1</code>, <code>S2</code>, <code>Night</code>, dll.)<br>'
    '- <b>Sekarang:</b> semua tipe shift hadir normal → <b><code>S</code></b> (seragam)<br><br>'

    '<b>4. WFS, WFA/1/2 WFA, AL/1/2 AL, UL/1/2 UL, DW, K, Late/1/2 UL dari durasi — tidak berubah.</b>'
    '<b>5. ➕ Empat Label Cuti Khusus Baru (Update Terbaru):</b><br>'
    '- <b>💍 HL</b> (Happy/Marry Leave): kolom <code>HL-Happy(Marry)-婚假(Day(s))</code> = 1 → label <code>HL</code>, ekspor warna kuning emas #FFE599<br>'
    '- <b>🤱 ML</b> (Maternity Leave): kolom <code>ML-MaternityLeave-产假(Day(s))</code> berisi nilai selain "--"/0/kosong → label <code>ML</code>, ekspor warna hijau #B6D7A8<br>'
    '- <b>👶 WML</b> (Wife Maternity Leave): kolom <code>WML-WifeMater-妻产假(Day(s))</code> = 1 → label <code>WML</code>, ekspor warna teal #A2C4C9<br>'
    '- <b>📝 OT</b> (Others): kolom <code>OT - Others - 其他(Day(s))</code> = 1 → label <code>OT</code>, ekspor warna abu-abu #D9D9D9<br>'
    '- Keempat label diperiksa <b>setelah WFA/½WFA</b> (langkah 9–12) dan <b>sebelum cek keterlambatan</b> (langkah 13–14)<br>'
    '- Bersifat <b>standalone</b> — karyawan dengan HL/ML/WML/OT tidak dikenai cek durasi keterlambatan<br><br>'
    '<b>6. 🖥️ Peningkatan UI Default View (Update Terbaru):</b><br>'
    '- <b>Light/Dark Mode</b>: kontras otomatis via CSS variables — teks, border, badge, dan background menyesuaikan tema OS<br>'
    '- <b>Tombol Upload</b>: digeser ke pojok kanan atas bersama tombol Logic, menggunakan layout kolom [5,1]<br>'
    '- <b>Tombol Back</b>: muncul di halaman rekap untuk kembali ke tabel riwayat periode<br>'
    '- <b>Tabel Riwayat</b>: kolom No. | Month | Periode | Upload Date | Created By | Aksi (Buka)<br><br>'
    '<br><br>'
    '<b>7. 📥 Fitur Export Multi-Periode (Update Terbaru):</b><br>'
    '- Tombol <b>📥 Export</b> ditambahkan di action bar (kanan atas, di antara Upload dan Logic)<br>'
    '- Saat diklik, muncul <b>panel month selector</b> — multiselect daftar periode yang sudah ada di DB<br>'
    '- Pengguna memilih satu atau lebih bulan, lalu klik <b>Download</b><br>'
    '- Hasil: satu file <code>.xlsx</code> dengan <b>sheet terpisah per periode</b> — tanpa perlu upload ulang<br>'
    '- Data diambil langsung dari database melalui <code>_get_all_daily_from_db()</code><br>'
    '- Fungsi internal <code>_populate_calendar_ws()</code> dipakai bersama oleh export reguler dan multi-periode<br>'
    '<b>8. 🖱️ Fix Dialog "Rincian Harian Karyawan" (Bug Fix):</b><br>'
    '- <b>Root cause:</b> <code>st.rerun()</code> eksplisit dipanggil setelah set <code>dialog_target = "detail"</code>, '
    'menyebabkan blok <code>else</code> me-reset state sebelum dialog sempat dirender<br>'
    '- <b>Fix:</b> Hapus <code>st.rerun()</code> dari blok seleksi baris — <code>on_select="rerun"</code> '
    'sudah men-trigger rerun otomatis, tidak perlu rerun kedua<br>'
    '- Sederhanakan kondisi: dialog dibuka untuk baris apapun yang dipilih, '
    'kecuali dialog employee yang sama sudah di-close secara eksplisit (<code>dialog_target == "closed"</code>)<br>'
    '- Hapus blok <code>else</code> yang me-reset <code>dialog_target = None</code> saat tidak ada baris dipilih, '
    'karena dialog mengelola lifecycle-nya sendiri setelah dibuka<br><br>'
    '<b>9. 🔴 Klasifikasi Baru H — Hari Libur Nasional (Tanggal Merah):</b><br>'
    '- Status <b>H</b> tidak dihasilkan oleh engine klasifikasi otomatis — hanya diterapkan via fitur <b>Bulk Correction</b><br>'
    '- <b>Alur kerja Bulk Correction:</b> pilih periode → pilih tanggal merah → pilih Rules (atau semua) → konfirmasi → semua record yang cocok di-update ke H dengan <code>is_manual_override = 1</code><br>'
    '- Ekspor kalender: sel H diberi warna <code style="background:#FF9999;padding:1px 6px;border-radius:3px;">H (#FF9999)</code><br>'
    '- Status H bersifat override penuh — menimpa apapun yang sebelumnya ada di kolom <code>status_klasifikasi</code><br>'
    '- Tombol <b>🔴 Tanggal Merah</b> tersedia di action bar kanan atas, dapat diakses tanpa membuka periode terlebih dahulu<br><br>'
    '</div>'

    '<div style="font-weight:700;color:#0f172a;margin-bottom:0.4rem;font-size:0.82rem;'
    'text-transform:uppercase;letter-spacing:0.06em;">✏️ Fitur Edit Manual</div>'

    '<div style="background:#f0fdf4;border-radius:8px;padding:0.7rem 1rem;margin-bottom:1.2rem;'
    'font-size:0.82rem;border-left:3px solid #22c55e;">'
    '<b>Cara akses:</b> Klik baris karyawan di tabel utama → buka expander '
    '<i>"✏️ Edit Data Karyawan &amp; Absensi Harian"</i> di bagian bawah dialog.<br><br>'

    '<b>1. Edit Data Karyawan</b><br>'
    '&nbsp;&nbsp;• <b>Nama Karyawan</b> — koreksi ejaan atau perubahan nama resmi<br>'
    '&nbsp;&nbsp;• <b>Rules / Departemen</b> — pindah divisi atau koreksi pengelompokan<br>'
    '&nbsp;&nbsp;• Perubahan langsung tersimpan ke database dan tercermin di rekap berikutnya<br><br>'

    '<b>2. Edit Absensi Harian (per baris)</b><br>'
    '&nbsp;&nbsp;• <b>Jam Masuk / Jam Keluar</b> — koreksi jam jika ada salah input atau missed punch<br>'
    '&nbsp;&nbsp;• <b>Klasifikasi</b> (dropdown) — override manual: pilih salah satu dari '
    '<code>S, Late, 1/2 UL, UL, AL, 1/2 AL, WFA, 1/2 WFA, WFS, DW, K, Off, HL, ML, WML, OT, None</code><br>'
    '&nbsp;&nbsp;• <b>Catatan / Keterangan</b> (kolom baru) — teks bebas, misal: '
    '"izin lisan", "sakit tanpa surat", "dinas luar kota"<br>'
    '&nbsp;&nbsp;• <b>Override</b> (checkbox) — centang untuk menandai baris yang '
    'ditetapkan manual; ditampilkan sebagai ✏️ di kolom detail<br><br>'

    '<b>Catatan penting:</b><br>'
    '&nbsp;&nbsp;• Edit hanya tersedia untuk data yang sudah tersimpan di database '
    '(terjadi otomatis setiap kali file Excel diupload)<br>'
    '&nbsp;&nbsp;• Override tidak mengubah logika engine — hanya menimpa nilai '
    '<code>status_klasifikasi</code> di DB; rekap otomatis mengikuti nilai DB<br>'
    '&nbsp;&nbsp;• Setelah menyimpan, tutup dan buka kembali dialog untuk melihat perubahan'
    '</div>'


    '<div style="font-weight:700;color:#0f172a;margin-bottom:0.4rem;font-size:0.82rem;'
    'text-transform:uppercase;letter-spacing:0.06em;">🔒 Semua Status Bersifat Standalone</div>'

    '<div style="background:#eff6ff;border-radius:8px;padding:0.6rem 1rem;margin-bottom:1.2rem;'
    'font-size:0.82rem;border-left:3px solid #3b82f6;">'
    'Setiap baris absensi menghasilkan tepat <b>satu status</b>:<br>'
    '- 📋 <b>S</b>: hadir tepat waktu → ekspor: <code style="background:#D9EAD3;padding:1px 6px;border-radius:3px;">S</code><br>'
    '- 🕐 <b>Late</b>: keterlambatan/pulang cepat 1-120 mnt → ekspor: <code style="background:#FFF2CC;padding:1px 6px;border-radius:3px;">L</code><br>'
    '- ⛔ <b>1/2 UL</b>: keterlambatan/pulang cepat &gt;120 mnt, atau UL kolom 0.5 → ekspor: <code style="background:#FCE5CD;padding:1px 6px;border-radius:3px;">0,5UL</code><br>'
    '- 📋 <b>UL</b>: Unpaid Leave penuh (kolom UL = 1) → ekspor: <code style="background:#B7E1CD;padding:1px 6px;border-radius:3px;">UL</code><br>'
    '- 🌴 <b>AL</b>: Annual Leave penuh (kolom AL = 1) → ekspor: <code style="background:#D9D2E9;padding:1px 6px;border-radius:3px;">AL</code><br>'
    '- 🌗 <b>1/2 AL</b>: Annual Leave setengah hari (kolom AL = 0.5) → ekspor: <code style="background:#EAD1DC;padding:1px 6px;border-radius:3px;">0,5AL</code><br>'
    '- 🏠 <b>WFA</b>: Work From Home penuh (kolom WFH = 1) → ekspor: <code style="background:#D0E4F7;padding:1px 6px;border-radius:3px;">WFA</code><br>'
    '- 🏡 <b>1/2 WFA</b>: Work From Home setengah hari (kolom WFH = 0.5) → ekspor: <code style="background:#C9DAF8;padding:1px 6px;border-radius:3px;">0,5WFA</code><br>'
    '- 📍 <b>WFS</b>: Work From Offsite → ekspor: <code style="background:#B4C7E7;padding:1px 6px;border-radius:3px;">WFS</code><br>'
    '- 💊 <b>K</b>: sakit dengan surat → ekspor: <code style="background:#F4CCCC;padding:1px 6px;border-radius:3px;">K</code><br>'
    '- 🚫 <b>DW</b>: tidak hadir → ekspor: <code style="background:#EA9999;padding:1px 6px;border-radius:3px;">DW</code><br>'
    '- 🏖️ <b>Off</b>: hari libur / tidak terjadwal → ekspor: <code style="background:#CFE2F3;padding:1px 6px;border-radius:3px;">OFF</code><br>'
    '- 💍 <b>HL</b>: cuti pernikahan → ekspor: <code style="background:#FFE599;padding:1px 6px;border-radius:3px;">HL</code><br>'
    '- 🤱 <b>ML</b>: cuti melahirkan → ekspor: <code style="background:#B6D7A8;padding:1px 6px;border-radius:3px;">ML</code><br>'
    '- 👶 <b>WML</b>: cuti istri melahirkan → ekspor: <code style="background:#A2C4C9;padding:1px 6px;border-radius:3px;">WML</code><br>'
    '- 📝 <b>OT</b>: cuti lainnya → ekspor: <code style="background:#D9D9D9;padding:1px 6px;border-radius:3px;">OT</code><br>'
    '- ❓ <b>None</b>: tidak memenuhi kondisi manapun — sel ekspor <b>kosong</b> (putih)'
    '</div>'

    '<div style="font-weight:700;color:#0f172a;margin-bottom:0.4rem;font-size:0.82rem;'
    'text-transform:uppercase;letter-spacing:0.06em;">⚠️ Pengecualian &amp; Catatan Penting</div>'

    '<div style="background:#fef9ec;border-radius:8px;padding:0.6rem 1rem;'
    'font-size:0.82rem;border-left:3px solid #f59e0b;">'
    '- 🎨 Background color hanya diterapkan pada sel data (baris 3+), bukan header<br>'
    '- 📍 WFS dicek <em>sebelum</em> skip-shift — jika att = "Normal (Offsite)", baris tetap diproses meski shift kosong<br>'
    '- ⏭️ Shift <code>Rest</code> / <code>Not scheduled</code> / <code>--</code> / kosong '
    '&rarr; dilewati engine klasifikasi. Jika att bukan "Normal (rest)" / "Normal (Offsite)" maka tampil sebagai <b>❓ None</b><br>'
    '- ❓ Baris <b>None</b> <em>tidak</em> masuk perhitungan metric — sel ekspor <b>kosong tanpa warna</b><br>'
    '- 🔍 K diperiksa <em>sebelum</em> DW agar sakit-dengan-surat tidak tertimpa absensi<br>'
    '- 🛡️ Karyawan dengan K / DW / AL / UL / WFA / 1/2 WFA / WFS <b>tidak dikenai</b> cek keterlambatan<br>'
    '- 📊 Kolom AL, UL, WFH menggunakan <code>0.5</code> untuk setengah hari; mendukung koma desimal ("0,5")<br>'
    '- 🗄️ DB menggunakan separator <code>|</code> (pipe) untuk menghindari konflik dengan "1/2"<br>'
    '- 📋 Kolom KTP pada ekspor kalender <b>dikosongkan</b> — dapat diisi manual jika diperlukan<br>'
    '- ✏️ Semua tipe shift hadir (S1, S2, Night, dll.) ditampilkan sebagai <code>S</code> di ekspor'
    '- 💍🤱👶📝 Karyawan dengan HL / ML / WML / OT <b>tidak dikenai</b> cek keterlambatan (Late/½UL)<br>'
    '- 🤱 ML dipicu oleh nilai <b>apapun selain "--"/0/kosong</b> — termasuk jumlah hari seperti "90" atau "60"<br>'
    '- 💍👶📝 HL, WML, OT dipicu hanya jika nilai kolom = <b>1</b> (tepat satu hari penuh)'
    '</div>'

    '</div>'
)

@st.dialog("⚠️ Data Periode Sudah Ada", width="small")
def show_override_confirm_dialog(periode_label: str, periode: str):
    st.markdown(
        f"Bulan **{periode_label}** sudah ada di database. "
        f"Apakah ingin **override** data lama?",
    )
    st.caption("Data lama akan di-soft-delete sebelum data baru disimpan.")
    st.markdown("<div style='margin:0.8rem 0 0.4rem'></div>", unsafe_allow_html=True)
    _dc1, _dc2 = st.columns(2)
    with _dc1:
        if st.button("✅ Ya, Override", type="primary", use_container_width=True, key="dlg_override_yes"):
            st.session_state._override_confirmed_for = periode
            st.session_state._show_override_confirm  = False
            st.rerun()
    with _dc2:
        if st.button("❌ Tidak, Batalkan", use_container_width=True, key="dlg_override_no"):
            st.session_state._show_override_confirm      = False
            st.session_state._pending_override_periode   = None
            st.session_state._pending_file_bytes         = None
            st.session_state._override_confirmed_for     = None
            st.session_state.show_upload_panel           = False
            st.toast("Upload dibatalkan. Data lama tidak diubah.", icon="ℹ️")
            st.rerun()




@st.dialog("📋 Logic Klasifikasi Absensi", width="large")
def show_logic_dialog():
    st.markdown(_LOGIC_HTML, unsafe_allow_html=True)


# ──────────────────────────────────────────────────────────────
# UI Utama
# ──────────────────────────────────────────────────────────────

st.markdown(
    '<div class="app-header">'
    '<div class="badge">📊 HR TOOLS</div>'
    '<h1>🗓️ Absensi Rekap Generator</h1>'
    '<p>📂 Upload file Excel absensi &rarr; 🔍 Hitung S / Late / UL / AL / WFA / WFS / DW per karyawan &rarr; 📥 Download hasil</p>'
    '</div>',
    unsafe_allow_html=True,
)

# ── Top Action Bar: rata kanan ───────────────────────────────
_spacer, _action_col = st.columns([1, 5], gap="medium")
with _action_col:
    _ab1, _ab2, _ab3, _ab4 = st.columns([3, 2, 1, 1], gap="small")
    with _ab1:
        if st.button(
            "📤 Upload File Absensi",
            use_container_width=True,
            type="primary",
            key="btn_upload_top",
        ):
            st.session_state.show_upload_panel = not st.session_state.get("show_upload_panel", False)
            st.session_state.show_export_panel = False
            st.session_state.show_h_panel      = False
            st.session_state.pop("_auto_periode", None)
            st.rerun()
    with _ab2:
        _export_active = st.session_state.get("show_export_panel", False)
        if st.button(
            "📥 Export",
            use_container_width=True,
            type="secondary",
            key="btn_export_top",
        ):
            st.session_state.show_export_panel = not _export_active
            st.session_state.show_upload_panel = False
            st.session_state.show_h_panel      = False
            st.rerun()
    with _ab3:
        _h_active = st.session_state.get("show_h_panel", False)
        if st.button(
            "🔴",
            use_container_width=True,
            type="primary" if _h_active else "secondary",
            key="btn_h_top",
            help="National Holiday — Bulk correction tanggal merah (hari libur nasional)",
        ):
            st.session_state.show_h_panel      = not _h_active
            st.session_state.show_upload_panel = False
            st.session_state.show_export_panel = False
            st.rerun()
    with _ab4:
        if st.button(
            "📋",
            use_container_width=True,
            type="secondary",
            key="btn_logic_top",
            help="Logic Klasifikasi Absensi — Lihat aturan & urutan prioritas klasifikasi",
        ):
            st.session_state.dialog_target = "logic"
            st.session_state.dialog_emp    = None
            st.rerun()

if st.session_state.dialog_target == "logic":
    st.session_state.dialog_target = None
    show_logic_dialog()

# ── Export Panel ──────────────────────────────────────────────
if st.session_state.get("show_export_panel", False):
    _exp_periodes = get_periodes()
    if not _exp_periodes:
        st.warning("⚠️ Belum ada periode tersimpan. Upload file Excel terlebih dahulu.")
        st.session_state.show_export_panel = False
    else:
        st.markdown(
            '<div class="action-panel">'
            '<div class="action-panel-title">📥 Export Data Absensi</div>'
            '<div class="action-panel-desc">'
            'Pilih satu atau beberapa bulan — setiap bulan menjadi sheet terpisah dalam satu file <code>.xlsx</code>.'
            '</div>',
            unsafe_allow_html=True,
        )
        _ex_col1, _ex_col2 = st.columns([3, 2], gap="medium")
        with _ex_col1:
            import datetime as _dt_ex
            _period_label_map = {}
            for _pp in _exp_periodes:
                try:
                    _period_label_map[_pp] = (
                        _dt_ex.datetime.strptime(_pp, "%Y-%m").strftime("%B %Y")
                        + f"  ·  {_pp}"
                    )
                except Exception:
                    _period_label_map[_pp] = _pp

            _sel_export = st.multiselect(
                label="📅 Pilih Bulan / Periode",
                options=_exp_periodes,
                format_func=lambda x: _period_label_map.get(x, x),
                placeholder="Klik untuk memilih periode...",
                key="export_period_select",
            )

        with _ex_col2:
            st.markdown("<div style='height:30px'></div>", unsafe_allow_html=True)
            if _sel_export:
                # Reset prepared state when selection changes
                if _sel_export != st.session_state._last_export_sel:
                    st.session_state.export_prepared = False
                    st.session_state._last_export_sel = _sel_export

                _export_label = (
                    "_".join(_sel_export)
                    if len(_sel_export) <= 3
                    else f"{_sel_export[0]}_sd_{_sel_export[-1]}"
                )

                if not st.session_state.export_prepared:
                    if st.button(
                        f"⚙️ Siapkan {len(_sel_export)} Periode",
                        type="primary",
                        use_container_width=True,
                        key="btn_prepare_export",
                    ):
                        st.session_state.export_prepared = True
                        st.rerun()
                    st.caption("Klik untuk menyiapkan file sebelum download.")
                else:
                    with st.spinner("⚙️ Menyiapkan file..."):
                        _export_bytes = export_multi_period_bytes(tuple(_sel_export))
                    st.download_button(
                        label=f"📥 Download {len(_sel_export)} Periode (.xlsx)",
                        data=_export_bytes,
                        file_name=f"Absensi_Export_{_export_label}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                        type="primary",
                        key="btn_download_export",
                    )
                    st.caption(
                        f"✅ {len(_sel_export)} sheet  ·  "
                        + "  |  ".join(
                            _period_label_map.get(p, p).split("  ·  ")[0]
                            for p in _sel_export
                        )
                    )
            else:
                st.markdown(
                    '<div style="background:#f1f5f9;border-radius:10px;padding:0.9rem 1.2rem;'
                    'text-align:center;color:#64748b;font-size:0.85rem;">'
                    '⬅️ Pilih periode terlebih dahulu</div>',
                    unsafe_allow_html=True,
                )


# ── Panel Bulk Correction H (Tanggal Merah) ──────────────────
if st.session_state.get("show_h_panel", False):
    _hp_periodes = get_periodes()
    if not _hp_periodes:
        st.warning("⚠️ Belum ada periode tersimpan. Upload file Excel terlebih dahulu.")
        st.session_state.show_h_panel = False
    else:
        st.markdown(
            '<div class="action-panel">'
            '<div class="action-panel-title">🔴 Bulk Correction — Hari Libur Nasional (H)</div>'
            '<div class="action-panel-desc">'
            'Pilih periode, tanggal merah, dan rules yang terdampak. '
            'Sistem akan mengubah status seluruh karyawan yang cocok menjadi <b>H</b> secara massal.'
            '</div>',
            unsafe_allow_html=True,
        )
        _hp1, _hp2, _hp3 = st.columns([1, 2, 2], gap="medium")

        with _hp1:
            _hp_sel_periode = st.selectbox(
                "📅 Periode",
                options=_hp_periodes,
                key="hp_periode_select",
            )

        with _hp2:
            _hp_all_dates = get_dates_in_periode(_hp_sel_periode)
            _hp_sel_dates = st.multiselect(
                "📆 Pilih Tanggal Merah",
                options=_hp_all_dates,
                placeholder="Klik untuk memilih tanggal...",
                key="hp_dates_select",
            )

        with _hp3:
            _hp_all_rules = get_rules_in_periode(_hp_sel_periode)
            _hp_sel_rules = st.multiselect(
                "🏷️ Filter Rules (kosong = semua)",
                options=_hp_all_rules,
                placeholder="Semua Rules",
                key="hp_rules_select",
            )

        if _hp_sel_dates:
            _hp_rules_display = (
                ", ".join(_hp_sel_rules) if _hp_sel_rules else "**semua rules**"
            )
            st.info(
                f"ℹ️ Akan mengubah status menjadi **H** pada **{len(_hp_sel_dates)} tanggal** "
                f"yang dipilih untuk {_hp_rules_display}."
            )
            _hpc1, _hpc2 = st.columns([1, 5])
            with _hpc1:
                if st.button(
                    "✅ Terapkan",
                    type="primary",
                    use_container_width=True,
                    key="btn_apply_h",
                ):
                    with st.spinner("⚙️ Menerapkan koreksi..."):
                        _hp_n = bulk_update_h(
                            _hp_sel_periode,
                            _hp_sel_dates,
                            _hp_sel_rules if _hp_sel_rules else None,
                        )
                    st.cache_data.clear()
                    st.success(
                        f"✅ **{_hp_n} record** berhasil diupdate ke **H** "
                        f"pada {len(_hp_sel_dates)} tanggal di periode {_hp_sel_periode}."
                    )
        else:
            st.markdown(
                '<div style="background:#f1f5f9;border-radius:10px;padding:0.8rem 1.2rem;'
                'color:#64748b;font-size:0.85rem;">'
                '⬅️ Pilih tanggal merah terlebih dahulu</div>',
                unsafe_allow_html=True,
            )
        st.markdown('</div>', unsafe_allow_html=True)


# ── Ambil daftar periode ──────────────────────────────────────
periodes_tersedia = get_periodes()

# ── Inisialisasi variabel default ────────────────────────────
uploaded        = None
periode_dipilih = "- Upload file baru -"
_NEW_PERIODE_SENTINEL = "- Upload file baru -"

# ── Panel Upload (muncul saat tombol Upload diklik) ───────────
if st.session_state.get("show_upload_panel", False):
    st.markdown(
        '<div style="background:#f8fafc;border:1px solid #e2e8f0;border-radius:12px;'
        'padding:1.4rem 1.6rem;margin-bottom:1.5rem;">'
        '<div style="font-weight:700;font-size:0.95rem;color:#0f172a;margin-bottom:1rem;">'
        '📂 Upload File Excel Absensi</div>',
        unsafe_allow_html=True,
    )
    _up1, _up2 = st.columns([3, 2], gap="medium")
    with _up1:
        uploaded = st.file_uploader(
            label="Pilih file Excel absensi (.xlsx / .xls)",
            type=["xlsx", "xls"],
            help="Sheet 'General statistics and attendan' harus ada.",
            key="main_uploader",
        )
    with _up2:
        if periodes_tersedia:
            st.markdown(
                '<div style="font-size:0.82rem;font-weight:600;color:#334155;'
                'margin-bottom:0.4rem;">🗄️ Atau pilih periode tersimpan:</div>',
                unsafe_allow_html=True,
            )
            periode_dipilih = st.selectbox(
                label="Periode",
                options=[_NEW_PERIODE_SENTINEL] + periodes_tersedia,
                label_visibility="collapsed",
                key="periode_select",
            )
    st.markdown('</div>', unsafe_allow_html=True)

# ── Handle auto-pilih periode dari tombol "Buka" di tabel ────
if st.session_state.get("_auto_periode") and periode_dipilih == _NEW_PERIODE_SENTINEL:
    periode_dipilih = st.session_state.pop("_auto_periode")

if periode_dipilih == _NEW_PERIODE_SENTINEL and st.session_state.get("current_periode"):
    periode_dipilih = st.session_state.current_periode

# ── Default View: Tabel Riwayat Periode ──────────────────────
if not st.session_state.get("show_upload_panel", False) and uploaded is None and periode_dipilih == _NEW_PERIODE_SENTINEL:
    import datetime as _dt_mod

    # ── Section header ────────────────────────────────────────
    st.markdown(
        '<div style="'
        'display:flex;align-items:center;gap:.6rem;'
        'margin-bottom:1.2rem;'
        '">'
        '<div style="'
        'width:8px;height:8px;border-radius:50%;'
        'background:#f59e0b;'
        'box-shadow:0 0 6px #f59e0b;'
        '"></div>'
        '<span style="'
        'font-size:.72rem;font-weight:700;'
        'color:var(--text-muted);'
        'text-transform:uppercase;letter-spacing:.1em;'
        '">Riwayat Periode Absensi</span>'
        '</div>',
        unsafe_allow_html=True,
    )



        # ── Table header (HTML) ───────────────────────────────
    if periodes_tersedia:

        _ACCENTS = [
            ("#6366f1", "#ede9fe", "#4c1d95"),
            ("#3b82f6", "#dbeafe", "#1e3a8a"),
            ("#0ea5e9", "#e0f2fe", "#0c4a6e"),
            ("#10b981", "#d1fae5", "#064e3b"),
            ("#f59e0b", "#fef3c7", "#451a03"),
            ("#ef4444", "#fee2e2", "#450a0a"),
            ("#a855f7", "#f3e8ff", "#3b0764"),
        ]

        st.markdown("""
<style>
.pt-head {
    display: grid;
    grid-template-columns: 48px 1fr 150px 180px 180px;
    background: var(--table-header-bg, #1e293b);
    border: 1px solid var(--border-color, #334155);
    border-radius: 12px 12px 0 0;
    padding: 0 1rem;
    margin-bottom: -2px;
    position: relative;
    z-index: 1;
}
.pt-head-cell {
    padding: .65rem .5rem;
    font-size: .72rem; font-weight: 700;
    color: var(--text-faint, #94a3b8);
    text-transform: uppercase; letter-spacing: .09em;
}
.pt-outer {
    display: grid;
    grid-template-columns: 1fr auto;
    gap: 0.75rem;
    align-items: center;
    margin-bottom: 0;
}
.pt-card-wrap {
    border: 1px solid var(--border-color, #334155);
    border-radius: 12px;
    overflow-x: auto;
    overflow-y: visible;
    margin-bottom: 0;
}
.pt-row {
    display: grid;
    grid-template-columns: 48px 1fr 150px 180px 180px;
    padding: 0 1rem;
    border-bottom: 1px solid var(--border-color, #334155);
    align-items: center;
    background: var(--bg-secondary, #1e293b);
    transition: background .13s;
    position: relative;
    min-height: 70px;
    box-sizing: border-box;
}
.pt-row:last-child { border-bottom: none; }
.pt-row::before {
    content: '';
    position: absolute;
    left: 0; top: 15%; bottom: 15%;
    width: 3px; border-radius: 0 3px 3px 0;
    background: var(--pt-accent, #6366f1);
    opacity: .5; transition: opacity .13s;
}
.pt-row:hover::before { opacity: 0.85; }
.pt-num {
    display: flex; align-items: center; justify-content: center;
}
.pt-num-circle {
    width: 32px; height: 32px; border-radius: 50%;
    display: flex; align-items: center; justify-content: center;
    font-size: .78rem; font-weight: 700;
    color: #fff;
}
.pt-cell-main { padding: .75rem .5rem; }
.pt-month-name { font-weight: 700; font-size: .92rem; color: var(--text-primary, #f1f5f9); }
.pt-year { font-size: .8rem; color: var(--text-faint, #64748b); margin-left: .35rem; font-weight: 400; }
.pt-sub { font-size: .7rem; color: var(--text-faint, #64748b); margin-top: .15rem; }
.pt-badge {
    display: inline-flex; align-items: center;
    padding: .25rem .85rem; border-radius: 20px;
    font-family: monospace; font-size: .74rem; font-weight: 700;
    letter-spacing: .04em; white-space: nowrap;
}
.pt-muted {
    font-size: .82rem;
    color: var(--text-faint, #64748b);
    padding: .75rem .5rem;
}
.pt-footer {
    padding: .6rem 1rem;
    border: 1px solid var(--border-color, #334155);
    border-top: none;
    border-radius: 0 0 12px 12px;
    background: var(--bg-secondary, #1e293b);
    font-size: .75rem;
    color: var(--text-faint, #64748b);
    display: flex; align-items: center; gap: .5rem;
    margin-bottom: 1rem;
}

/* ── Responsive: period table kolom tersembunyi di layar kecil ── */
@media (max-width: 800px) {
    .pt-head {
        grid-template-columns: 48px 1fr 140px !important;
    }
    .pt-row {
        grid-template-columns: 48px 1fr 140px !important;
    }
    .pt-head-cell:nth-child(4),
    .pt-head-cell:nth-child(5),
    .pt-row > div:nth-child(4),
    .pt-row > div:nth-child(5) {
        display: none !important;
    }
}

</style>
""", unsafe_allow_html=True)

        # ── Header ───────────────────────────────────────────
        _hcol, _hbtn_spacer = st.columns([10, 1], gap="small")
        with _hcol:
            st.markdown(
                '<div class="pt-head">'
                '<div class="pt-head-cell"></div>'
                '<div class="pt-head-cell">Month</div>'
                '<div class="pt-head-cell">Periode</div>'
                '<div class="pt-head-cell">Upload Date</div>'
                '<div class="pt-head-cell">Created By</div>'
                '</div>',
                unsafe_allow_html=True,
            )
        with _hbtn_spacer:
            st.markdown('<div style="min-height:1px"></div>', unsafe_allow_html=True)

        # ── Rows ─────────────────────────────────────────────
        st.markdown('<div class="pt-card-wrap">', unsafe_allow_html=True)

        for _i, _p in enumerate(periodes_tersedia):
            try:
                _dt_obj     = _dt_mod.datetime.strptime(_p, "%Y-%m")
                _month_name = _dt_obj.strftime("%B")
                _year_name  = _dt_obj.strftime("%Y")
            except Exception:
                _month_name = _p
                _year_name  = ""

            _accent, _badge_bg, _badge_fg = _ACCENTS[_i % len(_ACCENTS)]

            _col_row, _col_btn = st.columns([10, 1], gap="small")

            with _col_row:
                st.markdown(
                    f'<div class="pt-row" style="--pt-accent:{_accent};">'
                    f'  <div class="pt-num">'
                    f'    <div class="pt-num-circle" style="background:{_accent};">{_i + 1}</div>'
                    f'  </div>'
                    f'  <div class="pt-cell-main">'
                    f'    <div>'
                    f'      <span class="pt-month-name">{_month_name}</span>'
                    f'      <span class="pt-year">{_year_name}</span>'
                    f'    </div>'
                    f'    <div class="pt-sub">Periode absensi bulanan</div>'
                    f'  </div>'
                    f'  <div style="padding:.75rem .5rem;">'
                    f'    <span class="pt-badge" style="background:{_badge_bg};color:{_badge_fg};'
                    f'border:1px solid {_accent}55;">{_p}</span>'
                    f'  </div>'
                    f'  <div class="pt-muted">—</div>'
                    f'  <div class="pt-muted">—</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

            with _col_btn:
                if st.button(
                    "📂 Buka",
                    key=f"open_{_p}",
                    use_container_width=True,
                ):
                    st.session_state.show_upload_panel = True
                    st.session_state.show_export_panel = False
                    st.session_state["_auto_periode"]  = _p
                    st.rerun()

        st.markdown('</div>', unsafe_allow_html=True)

        # ── Footer ───────────────────────────────────────────
        st.markdown(
            f'<div class="pt-footer">'
            f'  <span style="width:7px;height:7px;border-radius:50%;'
            f'background:#6366f1;display:inline-block;flex-shrink:0;"></span>'
            f'  <b>{len(periodes_tersedia)} periode</b> tersimpan di database'
            f'  &nbsp;·&nbsp;'
            f'  Klik <b>📥 Export</b> untuk mengunduh tanpa membuka periode'
            f'</div>',
            unsafe_allow_html=True,
        )


    else:
        st.markdown(
            '<div style="'
            'text-align:center;padding:5rem 2rem;'
            '">'
            '<div style="font-size:4rem;margin-bottom:1.2rem;opacity:.35;">📁</div>'
            '<div style="'
            'font-size:1.05rem;font-weight:700;'
            'color:var(--text-muted);margin-bottom:.5rem;'
            '">Belum ada periode tersimpan</div>'
            '<div style="font-size:.85rem;color:var(--text-faint);">'
            'Klik <b>📤 Upload</b> di pojok kanan atas untuk mulai</div>'
            '</div>',
            unsafe_allow_html=True,
        )

    st.stop()


_NEW_PERIODE_SENTINEL = "- Upload file baru -"

if uploaded is not None or periode_dipilih != _NEW_PERIODE_SENTINEL:
    if uploaded is not None:
        # Baca file_bytes — simpan ke session agar tidak hilang saat rerun
        _raw_bytes = uploaded.read()
        if _raw_bytes:
            st.session_state._pending_file_bytes = _raw_bytes
        file_bytes = st.session_state._pending_file_bytes or b""

        if not file_bytes:
            st.error("❌ File tidak dapat dibaca. Coba upload ulang.")
            st.stop()

        with st.spinner("⚙️ Memproses data absensi..."):
            try:
                df_result, stats = process_file(file_bytes)
            except Exception as e:
                st.error(f"❌ Gagal memproses file: {e}")
                st.stop()

        # ── Konfirmasi Override (tampil jika periode sudah ada) ──────────
        # ── Konfirmasi Override (dialog modal) ──────────────────────────
        if st.session_state.get("_show_override_confirm"):
            _op = st.session_state._pending_override_periode
            try:
                _op_label = _dt.datetime.strptime(_op, "%Y-%m").strftime("%B %Y")
            except Exception:
                _op_label = _op
            show_override_confirm_dialog(_op_label, _op)
            st.stop()

        _periode  = None
        _save_ok  = True
        _save_done = False
        try:
            import io as _io, re as _re, pandas as _pd
            _buf = _io.BytesIO(file_bytes)
            df_raw = _pd.read_excel(
                _buf,
                sheet_name=_get_sheet_name(file_bytes),
                header=3,
                dtype={"Earliest": str, "Latest": str},
            ).rename(columns={"Unnamed: 0": "Time_Date", "Unnamed: 1": "Name", "Unnamed: 2": "Account"})
            for _val in df_raw["Time_Date"].astype(str):
                _m = _re.search(r'(\d{4})/(\d{2})/(\d{2})', _val)
                if _m:
                    _periode = f"{_m.group(1)}-{_m.group(2)}"
                    break
            if _periode is None:
                _periode = "unknown"

            df_raw = df_raw[df_raw["Account"].notna() & df_raw["Rules"].notna()]
            df_raw = df_raw[~df_raw["Account"].astype(str).str.strip().isin(["", "--"])]

            _k_sick_col       = _find_ksick_col(df_raw)
            _al_col           = _find_al_col(df_raw)
            _ul_col           = _find_ul_col(df_raw)
            _dur_late_col     = _find_duration_late_col(df_raw)
            _dur_early_col    = _find_duration_early_col(df_raw)
            _wfh_col          = _find_wfh_col(df_raw)
            _offsite_col      = _find_offsite_col(df_raw)
            _missed_punch_col = _find_missed_punch_col(df_raw)
            _hl_col           = _find_hl_col(df_raw)
            _ml_col           = _find_ml_col(df_raw)
            _wml_col          = _find_wml_col(df_raw)
            _ot_col           = _find_ot_col(df_raw)
            _rl_col           = _find_rl_col(df_raw)
            _pl_col           = _find_pl_col(df_raw)

            df_raw["_tipe_shift"] = df_raw["Shift"].apply(classify_shift_type)
            df_raw["_status_klasifikasi"] = df_raw.apply(
                lambda r: classify_str(
                    r["Earliest"], r["Shift"], r["Attendance results"],
                    latest_raw=r["Latest"],
                    leave_app=r.get("Leave & Overtime Application"),
                    absences_count=r.get("Number of absences(Count)"),
                    k_sick_count=r.get(_k_sick_col)         if _k_sick_col       else None,
                    al_count=r.get(_al_col)                  if _al_col           else None,
                    ul_count=r.get(_ul_col)                  if _ul_col           else None,
                    duration_late=r.get(_dur_late_col)       if _dur_late_col     else None,
                    duration_early=r.get(_dur_early_col)     if _dur_early_col    else None,
                    wfh_count=r.get(_wfh_col)                if _wfh_col          else None,
                    offsite_hour=r.get(_offsite_col)         if _offsite_col      else None,
                    missed_punch_count=r.get(_missed_punch_col) if _missed_punch_col else None,
                    hl_count=r.get(_hl_col)                  if _hl_col           else None,
                    ml_count=r.get(_ml_col)                  if _ml_col           else None,
                    wml_count=r.get(_wml_col)                if _wml_col          else None,
                    ot_count=r.get(_ot_col)                  if _ot_col           else None,
                    rl_count=r.get(_rl_col)                  if _rl_col           else None,
                    pl_count=r.get(_pl_col)                  if _pl_col           else None,
                ), axis=1,
            )

            # ── Cek apakah periode sudah ada di DB ──────────────────────
            _existing_periodes = get_periodes()
            if (
                _periode
                and _periode != "unknown"
                and _periode in _existing_periodes
                and st.session_state.get("_override_confirmed_for") != _periode
            ):
                # Periode sudah ada & belum dikonfirmasi → tahan, minta konfirmasi
                st.session_state._show_override_confirm    = True
                st.session_state._pending_override_periode = _periode
                _save_ok = False

            if _save_ok:
                # Jika override dikonfirmasi → soft delete dulu
                if st.session_state.get("_override_confirmed_for") == _periode:
                    soft_delete_periode(_periode)
                    st.session_state._override_confirmed_for = None

                _n_save_rows = len(df_raw)
                with st.spinner(
                    f"💾 Menyimpan {_n_save_rows:,} baris ke database "
                    f"(periode {_periode})… Mohon tunggu."
                ):
                    save_periode(df_raw, _periode)
                _save_done = True  # ← tandai sukses, navigasi di luar try

        except Exception as e:
            st.warning(f"⚠️ Gagal simpan ke database: {e}")

        # Jika perlu konfirmasi, tampilkan sekarang lalu stop
        if _save_done:
            st.session_state.current_periode     = _periode
            st.session_state._pending_file_bytes = None
            st.session_state.show_upload_panel   = False
            st.session_state["_auto_periode"]    = _periode
            st.cache_data.clear()
            st.rerun()
        elif not _save_ok:
            st.rerun()

    else:
        _periode = periode_dipilih
        st.session_state.current_periode = periode_dipilih

        df_raw_db = get_rekap(periode_dipilih)
        df_result = df_raw_db.rename(columns={
            "nama": "Nama", "account": "Account", "rules": "Rules",
            "normal": "S",
            "late": "Late", "half_ul": "1/2 UL",
            "ul_count": "UL",
            "half_al": "1/2 AL", "al": "AL",
            "wfa": "WFA",
            "half_wfa": "1/2 WFA",
            "wfs": "WFS",
            "dw": "DW", "k_sick": "K", "off_count": "Off",
            "hl": "HL", "ml": "ML", "wml": "WML", "ot": "OT", "rl": "RL", "h_count": "H", "pl_count": "PL",
        })
        for col in ["S", "Late", "1/2 UL", "UL", "AL", "1/2 AL",
                    "WFA", "1/2 WFA", "WFS", "DW", "K", "Off",
                    "HL", "ML", "WML", "OT", "RL","H","PL"]:
            if col not in df_result.columns:
                df_result[col] = 0
        file_bytes = None
        stats = {
            "total_rows": len(df_result),
            "classified": int(df_result[["S", "Late", "1/2 UL"]].sum().sum()),
            "skipped": 0,
            "employees": len(df_result),
            "dist": {},
        }
        df_result.insert(0, "No.", range(1, len(df_result) + 1))

    _back_col, _spacer2 = st.columns([1, 7])
    with _back_col:
        if st.button("← Back", key="btn_back_main", use_container_width=True, type="secondary"):
            st.session_state.show_upload_panel = False
            st.session_state.pop("_auto_periode", None)
            st.session_state.current_periode   = None
            # Clear cache agar saat buka periode lain tidak stale
            st.cache_data.clear()
            st.rerun()
    st.markdown("<div style='margin-bottom:0.5rem'></div>", unsafe_allow_html=True)

    total_s    = int(df_result["S"].sum())
    total_l    = int(df_result["Late"].sum())
    total_k    = int(df_result["1/2 UL"].sum())
    total_ul   = int(df_result["UL"].sum())
    total_al   = int(df_result["AL"].sum())
    total_hal  = int(df_result["1/2 AL"].sum())
    total_wfa  = int(df_result["WFA"].sum())
    total_hwfa = int(df_result["1/2 WFA"].sum())
    total_wfs  = int(df_result["WFS"].sum())
    total_dw   = int(df_result["DW"].sum())
    total_ks   = int(df_result["K"].sum())
    total_off  = int(df_result["Off"].sum())
    total_hl   = int(df_result["HL"].sum())  if "HL"  in df_result.columns else 0
    total_ml   = int(df_result["ML"].sum())  if "ML"  in df_result.columns else 0
    total_wml  = int(df_result["WML"].sum()) if "WML" in df_result.columns else 0
    total_ot   = int(df_result["OT"].sum())  if "OT"  in df_result.columns else 0
    total_rl   = int(df_result["RL"].sum()) if "RL" in df_result.columns else 0
    total_h    = int(df_result["H"].sum())  if "H"  in df_result.columns else 0
    total_pl   = int(df_result["PL"].sum()) if "PL" in df_result.columns else 0
    total_e    = stats["employees"]

    _section_label_style = (
        'style="font-size:.68rem;font-weight:700;color:var(--text-muted);'
        'text-transform:uppercase;letter-spacing:.1em;'
        'margin:1.4rem 0 .45rem;display:flex;align-items:center;gap:.5rem;"'
    )
    st.markdown(f"""
<div {_section_label_style}>
  <span style="width:6px;height:6px;border-radius:50%;background:#3b82f6;display:inline-block;flex-shrink:0;"></span>
  Attendance
</div>
<div class="metric-row" style="grid-template-columns: repeat(auto-fill, minmax(160px, 1fr));margin-top:0;">
  <div class="metric-card metric-shift">
    <div class="label"><span>📋</span> S (Shift)</div>
    <div class="value">{total_s:,}</div>
    <div class="sub">Hadir tepat / lebih awal</div>
  </div>
  <div class="metric-card metric-late">
    <div class="label"><span>🕐</span> Late</div>
    <div class="value">{total_l:,}</div>
    <div class="sub">Terlambat 1-120 mnt</div>
  </div>
  <div class="metric-card metric-k">
    <div class="label"><span>⛔</span> 1/2 UL</div>
    <div class="value">{total_k:,}</div>
    <div class="sub">Terlambat &gt;120 mnt / UL ½ hr</div>
  </div>
  <div class="metric-card metric-ul">
    <div class="label"><span>📋</span> UL</div>
    <div class="value">{total_ul:,}</div>
    <div class="sub">Unpaid Leave penuh</div>
  </div>
  <div class="metric-card metric-dw">
    <div class="label"><span>🚫</span> DW</div>
    <div class="value">{total_dw:,}</div>
    <div class="sub">Absence / Tidak hadir</div>
  </div>
  <div class="metric-card metric-total">
    <div class="label"><span>👥</span> Karyawan</div>
    <div class="value">{total_e:,}</div>
    <div class="sub">Total dalam periode</div>
  </div>
</div>

<div {_section_label_style}>
  <span style="width:6px;height:6px;border-radius:50%;background:#a855f7;display:inline-block;flex-shrink:0;"></span>
  Leave
</div>
<div class="metric-row" style="grid-template-columns: repeat(auto-fill, minmax(160px, 1fr));margin-top:0;">
  <div class="metric-card metric-ksick">
    <div class="label"><span>💊</span> K-Sick</div>
    <div class="value">{total_ks:,}</div>
    <div class="sub">Sakit dengan surat</div>
  </div>
  <div class="metric-card metric-al">
    <div class="label"><span>🌴</span> AL</div>
    <div class="value">{total_al:,}</div>
    <div class="sub">Annual Leave penuh</div>
  </div>
  <div class="metric-card metric-half-al">
    <div class="label"><span>🌗</span> 1/2 AL</div>
    <div class="value">{total_hal:,}</div>
    <div class="sub">Annual Leave setengah hari</div>
  </div>
  <div class="metric-card metric-wfa">
    <div class="label"><span>🏠</span> WFA</div>
    <div class="value">{total_wfa:,}</div>
    <div class="sub">Work From Home penuh</div>
  </div>
  <div class="metric-card metric-half-wfa">
    <div class="label"><span>🏡</span> 1/2 WFA</div>
    <div class="value">{total_hwfa:,}</div>
    <div class="sub">Work From Home ½ hari</div>
  </div>
  <div class="metric-card metric-wfs">
    <div class="label"><span>📍</span> WFS</div>
    <div class="value">{total_wfs:,}</div>
    <div class="sub">Work From Offsite</div>
  </div>
  <div class="metric-card metric-off">
    <div class="label"><span>🏖️</span> Off</div>
    <div class="value">{total_off:,}</div>
    <div class="sub">Rest / Not scheduled</div>
  </div>
</div>

<div {_section_label_style}>
  <span style="width:6px;height:6px;border-radius:50%;background:#f59e0b;display:inline-block;flex-shrink:0;"></span>
  Special Leave
</div>
<div class="metric-row" style="grid-template-columns: repeat(auto-fill, minmax(160px, 1fr));margin-top:0;">
  <div class="metric-card metric-hl">
    <div class="label"><span>💍</span> HL</div>
    <div class="value">{total_hl:,}</div>
    <div class="sub">Cuti Pernikahan</div>
  </div>
  <div class="metric-card metric-ml">
    <div class="label"><span>🤱</span> ML</div>
    <div class="value">{total_ml:,}</div>
    <div class="sub">Cuti Melahirkan</div>
  </div>
  <div class="metric-card metric-wml">
    <div class="label"><span>👶</span> WML</div>
    <div class="value">{total_wml:,}</div>
    <div class="sub">Cuti Istri Melahirkan</div>
  </div>
  <div class="metric-card metric-ot">
    <div class="label"><span>📝</span> OT</div>
    <div class="value">{total_ot:,}</div>
    <div class="sub">Cuti Lainnya</div>
  </div>
  <div class="metric-card metric-rl">
    <div class="label"><span>📅</span> RL</div>
    <div class="value">{total_rl:,}</div>
    <div class="sub">Roster Leave</div>
  </div>
  <div class="metric-card metric-h">
    <div class="label"><span>🔴</span> H</div>
    <div class="value">{total_h:,}</div>
    <div class="sub">Tanggal Merah</div>
  </div>
  <div class="metric-card metric-pl">
    <div class="label"><span>🪪</span> PL</div>
    <div class="value" style="color:#be185d;">{total_pl:,}</div>
    <div class="sub">Personal Leave (TKA)</div>
  </div>
</div>
""", unsafe_allow_html=True)

    st.markdown('<p class="section-title">📋 Hasil Summary per Karyawan</p>', unsafe_allow_html=True)

    fcol1, fcol2, fcol3 = st.columns([2, 2, 1])
    with fcol1:
        all_rules = sorted(df_result["Rules"].unique().tolist())
        sel_rules = st.multiselect("🏷️ Filter Rules", options=all_rules, placeholder="Semua Rules")
    with fcol2:
        search = st.text_input("🔍 Cari Nama / Account", placeholder="Ketik nama atau account...")
    with fcol3:
        show_late_only = st.checkbox("⚠️ Hanya Late/K/DW", value=False)

    # ── Active filter indicator ─────────────────────────────────────────
    _active_filters = []
    if sel_rules:
        _active_filters.append(f"🏷️ Rules: {', '.join(sel_rules)}")
    if search:
        _active_filters.append(f"🔍 \"{search}\"")
    if show_late_only:
        _active_filters.append("⚠️ Late / K / DW saja")

    if _active_filters:
        _tags_html = "".join(
            f'<span style="display:inline-flex;align-items:center;gap:.3rem;'
            f'background:var(--badge-bg);color:var(--badge-color);'
            f'border:1px solid rgba(29,78,216,.2);border-radius:20px;'
            f'padding:.18rem .65rem;font-size:.74rem;font-weight:600;'
            f'font-family:\'DM Mono\',monospace;white-space:nowrap;">'
            f'{f}</span>'
            for f in _active_filters
        )
        st.markdown(
            f'<div style="display:flex;align-items:center;gap:.5rem;'
            f'padding:.45rem .75rem;margin-bottom:.6rem;'
            f'background:var(--bg-secondary);border:1px solid var(--border-color);'
            f'border-radius:8px;flex-wrap:wrap;">'
            f'<span style="font-size:.72rem;color:var(--text-muted);font-weight:600;'
            f'text-transform:uppercase;letter-spacing:.06em;white-space:nowrap;">Filter aktif:</span>'
            f'{_tags_html}'
            f'</div>',
            unsafe_allow_html=True,
        )



    with st.expander("👁️ Tampilkan / Sembunyikan Kolom Kategori", expanded=False):
        st.markdown(
            '<div style="font-size:0.83rem;color:#64748b;margin-bottom:0.6rem;">'
            'Kolom <b>No., Nama, Account, Rules, S, Late, 1/2 UL, UL, DW</b> selalu tampil. '
            'Pilih kategori tambahan di bawah:</div>',
            unsafe_allow_html=True,
        )
        opt_cols_selected = []
        _r1 = st.columns(7)
        _r2 = st.columns(6)
        opt_col_ui = _r1 + _r2
        for i, (key, label, desc) in enumerate(OPTIONAL_COLS_DEF):
            with opt_col_ui[i]:
                checked = st.checkbox(label, value=True, help=desc, key=f"col_{key}")
                if checked:
                    opt_cols_selected.append(key)

    seen = set()
    visible_cols = []
    for c in CORE_COLS + opt_cols_selected:
        if c not in seen:
            seen.add(c)
            visible_cols.append(c)

    df_show = df_result.copy()
    if sel_rules:
        df_show = df_show[df_show["Rules"].isin(sel_rules)]
    if search:
        mask = (
            df_show["Nama"].str.contains(search, case=False, na=False) |
            df_show["Account"].str.contains(search, case=False, na=False)
        )
        df_show = df_show[mask]
    if show_late_only:
        df_show = df_show[
            (df_show["Late"] > 0) | (df_show["1/2 UL"] > 0) | (df_show["DW"] > 0)
        ]

    df_show = df_show.copy()
    df_show["No."] = range(1, len(df_show) + 1)

    hidden = [OPTIONAL_LABELS[k] for k in OPTIONAL_KEYS if k not in opt_cols_selected]
    db_source_note = "" if file_bytes is not None else "  |  🗄️ Data dari database"
    st.caption(
        "👆 Klik baris untuk melihat rincian harian" +
        db_source_note +
        (f"  |  Kolom tersembunyi: {', '.join(hidden)}" if hidden else "")
    )

    current_periode = st.session_state.get("current_periode") or _periode

    sel_event = st.dataframe(
        df_show[visible_cols],
        width="stretch",
        height=520,
        hide_index=True,
        on_select="rerun",
        selection_mode="single-row",
        column_config={k: v for k, v in COL_CONFIG_ALL.items() if k in visible_cols},
        key=f"df_summary_table_{st.session_state.df_key_suffix}",
    )

    current_periode = st.session_state.get("current_periode") or _periode
    sel_rows = sel_event.selection.rows if sel_event and sel_event.selection else []

    if sel_rows:
        idx = sel_rows[0]
        if idx < len(df_show):
            emp = df_show.iloc[idx]
            new_emp = {
                "account": emp["Account"],
                "nama"   : emp["Nama"],
                "rules"  : emp["Rules"],
                "periode": current_periode,
            }
            if not (
                st.session_state.dialog_target == "closed"
                and st.session_state.dialog_emp == new_emp
            ):
                st.session_state.dialog_target = "detail"
                st.session_state.dialog_emp    = new_emp

    if st.session_state.dialog_target == "logic":
        st.session_state.dialog_target = None
        show_logic_dialog()
    elif st.session_state.dialog_target == "detail" and st.session_state.dialog_emp:
        emp_s = st.session_state.dialog_emp
        show_daily_detail(
            account   = emp_s["account"],
            nama      = emp_s["nama"],
            rules     = emp_s["rules"],
            file_bytes= file_bytes,
            periode   = emp_s.get("periode"),
        )

    st.caption(
        f"📊 Menampilkan {len(df_show):,} dari {len(df_result):,} karyawan  |  "
        f"📄 Total baris diproses: {stats['total_rows']:,}  |  "
        f"✅ Diklasifikasikan: {stats['classified']:,}  |  "
        f"⏭️ Dilewati (Rest/dll): {stats['skipped']:,}"
    )

    st.markdown("---")

    time_range = ""
    if file_bytes is not None:
        try:
            buf_tr = io.BytesIO(file_bytes)
            raw_tr = pd.read_excel(buf_tr, sheet_name=_get_sheet_name(file_bytes),
                                   header=None, nrows=2)
            tr_text = str(raw_tr.iloc[1, 0])
            m_tr = re.search(r'Time Range[:\s]*([\d/\u2013\-\s]+)', tr_text)
            time_range = m_tr.group(1).strip() if m_tr else ""
        except Exception:
            time_range = ""

    fname = (
        "Absensi_Kalender_" + time_range.replace(" ", "_").replace("\u2013", "sd") + ".xlsx"
        if time_range else f"Absensi_Kalender_{current_periode or ''}.xlsx"
    )

    # df_daily_cal diambil di sini — akan di-refresh ulang setelah expander jika ada koreksi
    with st.spinner("⚙️ Menyiapkan data kalender..."):
        if current_periode:
            df_daily_cal = _get_all_daily_from_db(current_periode)
        elif file_bytes is not None:
            df_daily_cal = get_all_daily_for_calendar(file_bytes)
        else:
            df_daily_cal = pd.DataFrame()
    _df_daily_cal_periode = current_periode  # simpan untuk re-fetch setelah koreksi

    # ── Expander: Ringkasan Data Kosong (None) ────────────────────────────
    if not df_daily_cal.empty:
        # Melakukan rekonstruksi grid lengkap (karyawan x tanggal) untuk mencari sel yang benar-benar kosong di kalender
        _dates = sorted(df_daily_cal["Date"].dropna().unique())
        _emp_list = df_show[["Nama", "Account"]].drop_duplicates("Account").to_dict("records")

        # Map data klasifikasi yang ada
        _daily_map = {}
        _accounts_with_data: set = set()
        for _, row in df_daily_cal.iterrows():
            _acc = row["Account"]
            _accounts_with_data.add(_acc)
            if _acc not in _daily_map:
                _daily_map[_acc] = {}
            _daily_map[_acc][row["Date"]] = {
                "Shift":          row.get("Shift", ""),
                "Classification": row.get("Classification"),
                "AttResult":      row.get("AttResult", ""),
                "TipeShift":      row.get("TipeShift", ""),
                "JamMasuk":       row.get("JamMasuk", ""),
                "JamKeluar":      row.get("JamKeluar", ""),
            }

        _none_rows = []
        for _emp in _emp_list:
            _acc = _emp["Account"]
            _name = _emp["Nama"]
            for _d in _dates:
                _day_info = _daily_map.get(_acc, {}).get(_d)
                _record_exists = (_acc in _accounts_with_data) and (_day_info is not None)
                if _day_info is None:
                    _shift_t  = ""
                    _klas     = None
                    _att_res  = ""
                    _tipe_s   = ""
                    _jam_in   = ""
                    _jam_out  = ""
                else:
                    _shift_t  = _day_info["Shift"]
                    _klas     = _day_info["Classification"]
                    _att_res  = _day_info.get("AttResult", "")
                    _tipe_s   = _day_info.get("TipeShift", "")
                    _jam_in   = _day_info.get("JamMasuk", "")
                    _jam_out  = _day_info.get("JamKeluar", "")

                # Mengambil referensi langsung dari fungsi penentu tampilan sel kalender
                if not _get_cell_display(_shift_t, _klas):
                    _has_db_record = _day_info is not None
                    if _has_db_record:
                        _reason = determine_reason(
                            shift              = _shift_t,
                            att_result         = _att_res,
                            status_klasifikasi = _klas,
                            jam_masuk          = _jam_in,
                            jam_keluar         = _jam_out,
                        )
                    else:
                        _reason = "Tidak ada record di database"
                    _none_rows.append({
                        "Account":        _acc,
                        "Name":           _name,
                        "Date":           _d,
                        "Shift":          _shift_t,
                        "Classification": _klas,
                        "Reason":         _reason,
                        "HasRecord":      _has_db_record,
                    })
        _df_none = pd.DataFrame(_none_rows, columns=["Account", "Name", "Date", "Shift", "Classification", "Reason", "HasRecord"])

        # ── Step 1: Group _df_none per karyawan ──────────────────────────────
        # Struktur untuk outer table (ringkasan per karyawan)
        _none_grouped_rows = []
        _none_detail_map: dict[str, list[dict]] = {}  # account → list of {Date, Reason}

        if not _df_none.empty:
            for _acc_g, _grp in _df_none.groupby("Account", sort=False):
                _name_g   = _grp["Name"].iloc[0]
                _dates_g  = sorted(_grp["Date"].tolist())
                _n_g      = len(_dates_g)

                # Preview: maks 3 tanggal pertama + "…+N lagi" jika lebih
                _MAX_PREVIEW = 3
                if _n_g <= _MAX_PREVIEW:
                    _preview = ", ".join(_dates_g)
                else:
                    _preview = ", ".join(_dates_g[:_MAX_PREVIEW]) + f"  …+{_n_g - _MAX_PREVIEW} lagi"

                _none_grouped_rows.append({
                    "Account"       : _acc_g,
                    "Name"          : _name_g,
                    "n_dates"       : _n_g,
                    "dates_preview" : _preview,
                })

                # Detail map: list {Date, Reason} untuk sub-tabel
                _none_detail_map[_acc_g] = [
                    {"Date": row["Date"], "Reason": row["Reason"], "HasRecord": bool(row.get("HasRecord", True))}
                    for _, row in _grp.sort_values("Date").iterrows()
                ]

        # Sort by n_dates descending (karyawan dengan paling banyak kosong di atas)
        _df_none_grouped = (
            pd.DataFrame(_none_grouped_rows)
            .sort_values("n_dates", ascending=False)
            .reset_index(drop=True)
            if _none_grouped_rows else pd.DataFrame(
                columns=["Account", "Name", "n_dates", "dates_preview"]
            )
        )


        _n_emp_none  = len(_df_none_grouped)
        _n_days_none = len(_df_none)
        _none_label  = (
            f"⚠️ Data Kosong / Tidak Terklasifikasi — "
            f"{_n_emp_none} karyawan · {_n_days_none} hari"
        )

        with st.expander(_none_label, expanded=False):
            if _df_none.empty:
                st.success("✅ Semua data sudah terklasifikasi. Tidak ada baris kosong.")
            else:
                st.markdown(
                    '<div style="font-size:0.82rem;color:#64748b;margin-bottom:0.8rem;">'
                    'Klik nama karyawan untuk melihat detail per tanggal. '
                    'Sel kosong = tidak masuk klasifikasi manapun → tampil '
                    '<b>putih</b> di kalender.'
                    '</div>',
                    unsafe_allow_html=True,
                )

                # ── CSS sub-tabel ────────────────────────────────────────
                st.markdown("""
<style>
.nt-sub-wrap {
    border: 1px solid var(--border-color, #e2e8f0);
    border-top: 2px solid rgba(245,158,11,.35);
    border-radius: 0 0 8px 8px;
    overflow: hidden;
    margin-top: 0.3rem;
}
.nt-sub-head-row { background: #fffbeb; }
.nt-sub-row-even { background: #f8fafc; }
.nt-sub-row-odd  { background: #ffffff; }
.nt-sub-date {
    padding: .4rem .8rem; font-family: monospace;
    font-size: .78rem; color: #475569; white-space: nowrap; width: 130px;
}
.nt-sub-reason { padding: .4rem .8rem; }
@media (prefers-color-scheme: dark) {
    .nt-sub-head-row { background: #1c1917 !important; }
    .nt-sub-row-even { background: #1e293b !important; }
    .nt-sub-row-odd  { background: #0f172a !important; }
    .nt-sub-date     { color: #94a3b8 !important; }
}
</style>
""", unsafe_allow_html=True)

                # ── Bersihkan sisa toggle keys lama (jika ada) ──────────
                for _stale_k in [
                    k for k in list(st.session_state)
                    if k.startswith("_nt_exp_")
                ]:
                    del st.session_state[_stale_k]

                # ── Helper badge reason ──────────────────────────────────
                def _reason_badge(reason: str) -> str:
                    r = str(reason)
                    if "dilewati" in r:
                        bg, fg = "#dbeafe", "#1e40af"
                    elif "tidak tercatat" in r:
                        bg, fg = "#fee2e2", "#991b1b"
                    elif "tidak dikenali" in r:
                        bg, fg = "#fef3c7", "#92400e"
                    else:
                        bg, fg = "#f1f5f9", "#475569"
                    return (
                        f'<span style="display:inline-flex;align-items:center;'
                        f'background:{bg};color:{fg};padding:.15rem .6rem;'
                        f'border-radius:4px;font-size:.72rem;font-weight:500;">'
                        f'{reason}</span>'
                    )

                _ALL_KLASIFIKASI_OPTS = [
                    "", "S", "Late", "1/2 UL", "UL", "AL", "1/2 AL",
                    "WFA", "1/2 WFA", "WFS", "DW", "K", "Off",
                    "HL", "ML", "WML", "OT", "RL", "H",
                ]

                # ── Rows: st.expander per karyawan ──────────────────────
                _all_none_edits: list[dict] = []  # kumpulkan semua edit lintas karyawan

                for _ri, _er in _df_none_grouped.iterrows():
                    _acc_e  = _er["Account"]
                    _name_e = _er["Name"]
                    _n_e    = int(_er["n_dates"])

                    _exp_label = (
                        f"**{_ri + 1}.** {_name_e}"
                        f"  ·  `{_acc_e}`"
                        f"  ·  📅 **{_n_e} hari kosong**"
                    )

                    with st.expander(_exp_label, expanded=False):
                        _detail_rows = _none_detail_map.get(_acc_e, [])
                        if not _detail_rows:
                            st.markdown(
                                '<div style="padding:.55rem 1rem;background:#f8fafc;'
                                'border-radius:6px;font-size:.82rem;color:#94a3b8;">'
                                'Tidak ada detail tersedia.</div>',
                                unsafe_allow_html=True,
                            )
                            continue

                        # Bangun DataFrame untuk data_editor
                        # Ambil remarks yang sudah tersimpan dari df_daily_cal
                        _remarks_map: dict[str, str] = {}
                        if not df_daily_cal.empty and "Remarks" in df_daily_cal.columns:
                            _emp_daily = df_daily_cal[df_daily_cal["Account"] == _acc_e]
                            _remarks_map = dict(
                                zip(_emp_daily["Date"], _emp_daily["Remarks"])
                            )

                        _editor_rows = []
                        for _sd in _detail_rows:
                            _editor_rows.append({
                                "Tanggal"   : _sd["Date"],
                                "Alasan"    : _sd["Reason"],
                                "Edit"      : "" if _sd.get("HasRecord", True) else "—",
                                "Remarks"   : _remarks_map.get(_sd["Date"], "") if _sd.get("HasRecord", True) else "",
                                "HasRecord" : _sd.get("HasRecord", True),
                            })
                        _editor_df = pd.DataFrame(_editor_rows)

                        _edited = st.data_editor(
                            _editor_df,
                            key=f"_none_editor_{_acc_e}",
                            use_container_width=True,
                            hide_index=True,
                            height=min(60 + len(_editor_df) * 35, 400),
                            column_config={
                                "Tanggal": st.column_config.TextColumn(
                                    "📅 Tanggal",
                                    disabled=True,
                                    width="small",
                                ),
                                "Alasan": st.column_config.TextColumn(
                                    "💬 Alasan",
                                    disabled=True,
                                    width="large",
                                ),
                                "Edit": st.column_config.SelectboxColumn(
                                    "✏️ Edit Klasifikasi",
                                    options=_ALL_KLASIFIKASI_OPTS,
                                    required=False,
                                    width="medium",
                                    help="Pilih klasifikasi yang seharusnya untuk baris ini",
                                ),
                                "Remarks": st.column_config.TextColumn(
                                    "📝 Remarks",
                                    width="large",
                                    help="Catatan bebas — akan muncul sebagai Comment di ekspor Excel",
                                ),
                                "HasRecord": None,  # sembunyikan kolom internal
                            },
                        )

                        # Kumpulkan baris yang ada perubahan (Edit tidak kosong atau Remarks terisi)
                        for _, _erow in _edited.iterrows():    
                            _raw_edit    = _erow.get("Edit")
                            _edit_val    = str(_raw_edit).strip() if (_raw_edit is not None and str(_raw_edit).strip() not in ("", "nan", "None", "—")) else ""
                            _remarks_val = str(_erow.get("Remarks", "") or "").strip()
                            if _edit_val or _remarks_val:
                                _all_none_edits.append({
                                    "account"    : _acc_e,
                                    "tanggal"    : _erow["Tanggal"],
                                    "status"     : _edit_val if _edit_val else None,
                                    "remarks"    : _remarks_val,
                                    "has_record" : bool(_erow.get("HasRecord", True)),
                                    "periode"    : current_periode or "",
                                })

                # ── Tombol Simpan semua koreksi ─────────────────────────
                _has_edits = any(
                    (e["status"] or e["remarks"]) for e in _all_none_edits
                )
                _save_col, _ = st.columns([1, 4])
                with _save_col:
                    if st.button(
                        "💾 Simpan Koreksi",
                        type="primary",
                        use_container_width=True,
                        key="btn_save_none_corrections",
                        disabled=not _has_edits,
                    ):
                        _n_saved = bulk_update_none_corrections(_all_none_edits)
                        st.cache_data.clear()
                        st.session_state["_force_refresh_cal"] = True
                        _edit_summary = ", ".join(
                            f"{e['account']}@{e['tanggal']}→{e['status'] or '(remarks only)'}"
                            for e in _all_none_edits
                        )
                        st.success(
                            f"✅ **{_n_saved} baris** berhasil dikoreksi. "
                            f"Detail: {_edit_summary}"
                        )
                        st.rerun()

                st.caption(
                    f"Total {_n_days_none} hari kosong "
                    f"dari {_n_emp_none} karyawan."
                    "  ·  Isi kolom **Edit** untuk mengubah klasifikasi, "
                    "**Remarks** untuk catatan yang akan muncul di ekspor Excel."
                )

    else:
        with st.expander("⚠️ Data Kosong / Tidak Terklasifikasi", expanded=False):
            st.info("Data kalender belum tersedia untuk diperiksa.")

        # ── Download Buttons — dirender setelah Data Kosong agar pakai data terbaru ──
    # Re-fetch df_daily_cal jika ada koreksi yang baru disimpan
    if st.session_state.pop("_force_refresh_cal", False):
        if current_periode:
            df_daily_cal = _get_all_daily_from_db(current_periode)
        elif file_bytes is not None:
            df_daily_cal = get_all_daily_for_calendar(file_bytes)
        else:
            df_daily_cal = pd.DataFrame()

    dcol1, dcol2 = st.columns([1, 1])

    with dcol1:
        xlsx_bytes = to_excel_calendar_bytes(df_daily_cal, df_result, time_range or current_periode or "")
        st.download_button(
            label="📥 Download Kalender Harian (.xlsx)",
            data=xlsx_bytes,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )

    with dcol2:
        if len(df_show) < len(df_result):
            visible_accs  = set(df_show["Account"].tolist())
            df_daily_filt = df_daily_cal[df_daily_cal["Account"].isin(visible_accs)] if not df_daily_cal.empty else df_daily_cal
            xlsx_filtered = to_excel_calendar_bytes(df_daily_filt, df_show, time_range or current_periode or "")
            st.download_button(
                label="🔽 Download Hasil Filter (.xlsx)",
                data=xlsx_filtered,
                file_name=f"Filter_{fname}",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
            )

    with st.expander("📊 Ringkasan per Rules"):
        grp = df_result.groupby("Rules").agg(
            Karyawan=("Account", "count"),
            S=("S", "sum"),
            Late=("Late", "sum"),
            **{"1/2 UL": ("1/2 UL", "sum")},
            UL=("UL", "sum"),
            DW=("DW", "sum"),
            K=("K", "sum"),
            AL=("AL", "sum"),
            WFA=("WFA", "sum"),
            **{"1/2 WFA": ("1/2 WFA", "sum")},
            WFS=("WFS", "sum"),
            Off=("Off", "sum"),
        ).reset_index().sort_values("1/2 UL", ascending=False)
        total_absen = grp["S"] + grp["Late"] + grp["1/2 UL"]
        grp["Late Rate"]    = (grp["Late"]   / total_absen.replace(0, 1) * 100).round(1)
        grp["1/2 UL Rate"]  = (grp["1/2 UL"] / total_absen.replace(0, 1) * 100).round(1)
        st.dataframe(
            grp,
            width="stretch",
            hide_index=True,
            column_config={
                "Rules"      : st.column_config.TextColumn("🏷️ Rules"),
                "Karyawan"   : st.column_config.NumberColumn("👥 Karyawan", format="%d"),
                "S"          : st.column_config.NumberColumn("📋 S (Shift)", format="%d"),
                "Late"       : st.column_config.NumberColumn("🕐 Late", format="%d"),
                "1/2 UL"     : st.column_config.NumberColumn("⛔ 1/2 UL", format="%d"),
                "UL"         : st.column_config.NumberColumn("📋 UL", format="%d"),
                "DW"         : st.column_config.NumberColumn("🚫 DW", format="%d"),
                "K"          : st.column_config.NumberColumn("💊 K", format="%d"),
                "AL"         : st.column_config.NumberColumn("🌴 AL", format="%d"),
                "WFA"        : st.column_config.NumberColumn("🏠 WFA", format="%d"),
                "1/2 WFA"    : st.column_config.NumberColumn("🏡 1/2 WFA", format="%d"),
                "WFS"        : st.column_config.NumberColumn("📍 WFS", format="%d"),
                "Off"        : st.column_config.NumberColumn("🏖️ Off", format="%d"),
                "Late Rate"  : st.column_config.NumberColumn("📈 % Late", format="%.1f%%"),
                "1/2 UL Rate": st.column_config.NumberColumn("📈 % 1/2 UL", format="%.1f%%"),
            },
        )

